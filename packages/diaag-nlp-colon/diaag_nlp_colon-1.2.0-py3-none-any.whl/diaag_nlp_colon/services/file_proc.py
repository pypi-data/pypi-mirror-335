import os
import pandas as pd
from diaag_nlp_colon.config.colon import brat_label_configs


# Functions to handle file processing

# Reads brat .ann files and report .txt files into memory
# Returns:
# dict of file_id -> brat report text (brat_dict)
# dict of file_id -> raw report text (report_dict)
# list of (raw report text, filename) tuples (report_list)
def read_report_files(paths):
    brat_dict = {}
    report_dict = {}
    report_list = []
    print('\nReading reports...')
    for path, old_reports in paths.items():
        print(path)
        # Read every file in directory
        for filename in os.listdir(path):
            # 2 reports without matches
            # if '1381595' in filename or '4594438' in filename:
            #     continue
            # secret duplicate col report
            # if '5853187_Colonoscopy' in filename:
            #     continue
            try:
                with open((path + filename), 'r', encoding='utf-8') as f:
                    # ignore hidden files
                    if filename.startswith('.'):
                        continue
                    f_str = f.read()

                    file_id = filename[3:10]

                    if '.ann' in filename:
                        if file_id in brat_dict:
                            print('Duplicate ann file:', filename)
                        else:
                            brat_dict[file_id] = f_str
                    elif '.txt' in filename:
                        # Note: some brat files have escaped newline characters
                        # for some reason the annotations count those chars
                        # so I needed to add two characters as a placeholder
                        if old_reports:
                            if '\\n' in f_str:
                                f_str = f_str.replace('\\n', '  ')
                            else:
                                f_str = f_str.replace('\n', '  ')
                        if file_id in report_dict:
                            print('Duplicate txt file:', filename)
                        else:
                            report_dict[file_id] = f_str
                            report_list.append((f_str, {'filename': filename}))
            except UnicodeDecodeError:
                print('encoding error:', filename)

    return brat_dict, report_dict, report_list


# Turn Brat annotations into spacy entity labels
# Final result needs to be list of spacy-friendly training data, e.g:
#     (
#        "Two large polypoid fragments, 1 x 0.3 x 0.2 cm and 0.9 x 0.7 x 0.4 cm",
#        {"entities": [(31, 47, LABEL), (52, 60, LABEL)]},
#    )
def generate_path_dataset(brat_data, report_data):
    brat_labels = brat_label_configs.BRAT_LABELS['path']
    with_semi = 0
    spacy_dataset = []
    full_dataset = {}
    cyt_dys_counts = {}
    location_counts = {}
    location_other_counts = {}
    hg_dys_counts = {}
    quantity_counts = {}
    # make list of brat entity annotations (lines that start with T1, T2, ...)
    for filename in brat_data.keys():
        sample_ents = {'entities': []}
        full = brat_data[filename]
        lines = full.split('\n')
        for ann in lines:
            if ann.startswith('T'):
                # annotations are tab-separated (entID, name w/ offsets, text)
                #                 print(filename)
                #                 print(ann)
                ann = ann.split('\t')
                # only add some annotations if we're testing trained model
                if ann[1].startswith('path-sample'):
                    ent_label = brat_labels['path-sample']
                elif ann[1].startswith('path-size'):
                    ent_label = brat_labels['path-size']
                elif ann[1].startswith('path-quantity'):
                    ent_label = brat_labels['path-quantity']
                    ann_val = ann[2]
                    if ann_val not in quantity_counts:
                        quantity_counts[ann_val] = 1
                    else:
                        quantity_counts[ann_val] += 1
                elif ann[1].startswith('path-location'):
                    ent_label = brat_labels['path-location']
                    # this is just for looking into location-other values
                    exact_label = ann[1].split(' ')[0]
                    ann_val = ann[2]
                    if exact_label == 'path-location-other':
                        if ann_val not in location_other_counts:
                            location_other_counts[ann_val] = 1
                        else:
                            location_other_counts[ann_val] += 1
                    # track all locations
                    if ann_val not in location_counts:
                        location_counts[ann_val] = 1
                    else:
                        location_counts[ann_val] += 1
                elif ann[1].startswith('path-histology-cytologic'):
                    ent_label = brat_labels['path-histology-cytologic-dysplasia']
                    ann_val = ann[2]
                    if ann_val not in cyt_dys_counts:
                        cyt_dys_counts[ann_val] = 1
                    else:
                        cyt_dys_counts[ann_val] += 1
                elif ann[1].startswith('path-histology'):
                    ent_label = brat_labels['path-histology']
                elif ann[1].startswith('path-high-grade-dysplasia'):
                    ent_label = brat_labels['path-high-grade-dysplasia']
                    ann_val = ann[2]
                    if ann_val not in hg_dys_counts:
                        hg_dys_counts[ann_val] = 1
                    else:
                        hg_dys_counts[ann_val] += 1
                else:
                    continue
                # ignore discontinuous annotations for now
                ann_label = ann[1]
                if ';' in ann_label:
                    with_semi += 1
                    continue
                ann = ann[1].split()
                sample_ents['entities'].append((int(ann[1]), int(ann[2]), ent_label))
        report_data_tup = (report_data[filename], sample_ents)
        spacy_dataset.append(report_data_tup)
        if filename in full_dataset:
            print('duplicate report?', filename)
        else:
            full_dataset[filename] = report_data_tup

    # write_label_counts('path', [
    #     ('location', location_counts),
    #     ('location (other)', location_other_counts),
    #     ('high-grade dysplasia', hg_dys_counts),
    #     ('cytologic dysplasia', cyt_dys_counts),
    #     ('quantity', quantity_counts)
    # ])

    return spacy_dataset, full_dataset


# Turn Brat annotations into spacy entity labels
# Final result needs to be list of spacy-friendly training data, e.g:
#     (
#        "Two large polypoid fragments, 1 x 0.3 x 0.2 cm and 0.9 x 0.7 x 0.4 cm",
#        {"entities": [(31, 47, LABEL), (52, 60, LABEL)]},
#    )
def generate_col_dataset(brat_data, report_data):
    brat_labels = brat_label_configs.BRAT_LABELS['col']
    with_semi = 0
    spacy_dataset = []
    full_dataset = {}
    location_other_counts = {}
    location_counts = {}
    procedure_counts = {}
    size_nonspec_counts = {}
    morph_counts = {}
    quantity_counts = {}
    no_rel = 0
    # make list of brat entity annotations (lines that start with T1, T2, ...)
    for filename in brat_data.keys():
        sample_ents = {'entities': []}
        full = brat_data[filename]
        lines = full.split('\n')
        contains_relation = False
        for ann in lines:
            if ann.startswith('R'):
                contains_relation = True
            if ann.startswith('T'):
                # annotations are tab-separated (entID, name w/ offsets, text)
                ann = ann.split('\t')
                if ann[1].startswith('size-measurement'):
                    ent_label = brat_labels['size-measurement']
                elif ann[1].startswith('size-nonspecific'):
                    ent_label = brat_labels['size-nonspecific']
                    ann_val = ann[2]
                    if ann_val not in size_nonspec_counts:
                        size_nonspec_counts[ann_val] = 1
                    else:
                        size_nonspec_counts[ann_val] += 1
                elif ann[1].startswith('finding-polyp-quantity'):
                    ent_label = brat_labels['finding-polyp-quantity']
                    ann_val = ann[2]
                    if ann_val not in quantity_counts:
                        quantity_counts[ann_val] = 1
                    else:
                        quantity_counts[ann_val] += 1
                elif ann[1].startswith('finding-polyp'):
                    ent_label = brat_labels['finding-polyp']
                elif ann[1].startswith('location'):
                    ent_label = brat_labels['location']
                    # this is just for looking into location-other values
                    exact_label = ann[1].split(' ')[0]
                    ann_val = ann[2]
                    if exact_label == 'location-other':
                        if ann_val not in location_other_counts:
                            location_other_counts[ann_val] = 1
                        else:
                            location_other_counts[ann_val] += 1
                    # track all locations
                    if ann_val not in location_counts:
                        location_counts[ann_val] = 1
                    else:
                        location_counts[ann_val] += 1
                elif ann[1].startswith('morphology'):
                    ent_label = brat_labels['morphology']
                    ann_val = ann[2]
                    if ann_val not in morph_counts:
                        morph_counts[ann_val] = 1
                    else:
                        morph_counts[ann_val] += 1
                else:
                    continue
                # ignore discontinuous annotations for now
                ann_label = ann[1]
                if ';' in ann_label:
                    with_semi += 1
                    continue
                ann = ann[1].split()
                # avoid overlapping ents
                start = int(ann[1])
                end = int(ann[2])
                overlapping = False
                for ex in sample_ents['entities']:
                    ex_start = int(ex[0])
                    ex_end = int(ex[1])
                    if (start >= ex_start) and (start < ex_end):
                        overlapping = True
                        break
                if overlapping:
                    # print('skipping overlapping annotation')
                    continue
                sample_ents['entities'].append((int(ann[1]), int(ann[2]), ent_label))
        # if not contains_relation:
        #     no_rel += 1
        #     continue
        report_data_tup = (report_data[filename], sample_ents)
        spacy_dataset.append(report_data_tup)
        if filename in full_dataset:
            print('duplicate report?', filename)
        else:
            full_dataset[filename] = report_data_tup

    # write_label_counts('col', [
    #     ('location', location_counts),
    #     ('location (other)', location_other_counts),
    #     ('morphology', morph_counts),
    #     ('size (nonspecific)', size_nonspec_counts),
    #     ('quantity', quantity_counts)
    # ])
    # print('{} reports without relations'.format(no_rel))

    return spacy_dataset, full_dataset


# Helper function to write summary of Brat annotations values to excel sheet for review
def write_label_counts(report_type, label_counts):
    label_dfs = []
    for label in label_counts:
        label_name, count_dict = label
        sorted_dict = {k: v for k, v in sorted(count_dict.items(), key=lambda item: item[1], reverse=True)}
        df = pd.DataFrame.from_dict(sorted_dict, orient='index')
        df.reset_index(level=0, inplace=True)
        df.columns = ['Term', 'Count']
        label_dfs.append((label_name, df))

    with pd.ExcelWriter('../test_reports/dataset_summary/{}_feature_values.xlsx'.format(report_type)) as writer:
        for label in label_dfs:
            label_name, df = label
            df.to_excel(writer, sheet_name=label_name, index=False)


# courtesy of stackoverflow :')
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

