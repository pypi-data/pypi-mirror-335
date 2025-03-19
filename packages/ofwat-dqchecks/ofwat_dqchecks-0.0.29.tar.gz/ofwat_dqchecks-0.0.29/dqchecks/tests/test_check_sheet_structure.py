"""
Test check_sheet_structure function in panacea
"""

import pytest
from openpyxl import Workbook

from dqchecks.panacea import check_sheet_structure

@pytest.fixture
def sheet1():
    """Fixture for creating a sample sheet1."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    sheet['A1'] = 'Name'
    sheet['B1'] = 'Age'
    sheet['A2'] = 'Alice'
    sheet['B2'] = 30
    sheet['A3'] = 'Bob'
    sheet['B3'] = 25
    return sheet


@pytest.fixture
def sheet2():
    """Fixture for creating a sample sheet2."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Sheet2'
    sheet['A1'] = 'Name'
    sheet['B1'] = 'Age'
    sheet['A2'] = 'Alice'
    sheet['B2'] = 30
    sheet['A3'] = 'Bob'
    sheet['B3'] = 25
    return sheet

@pytest.fixture
def sheet_with_different_headers():
    """Fixture for creating a sheet with different column headers."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    sheet['A1'] = 'First Name'
    sheet['B1'] = 'Age'
    sheet['A2'] = 'Alice'
    sheet['B2'] = 30
    sheet['A3'] = 'Bob'
    sheet['B3'] = 25
    return sheet


@pytest.fixture
def empty_sheet():
    """Fixture for creating an empty sheet."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    return sheet

# pylint: disable=W0621
def test_check_sheet_structure_equal(sheet1, sheet2):
    """Test that two identical sheets return True with no differences."""
    result = check_sheet_structure(sheet1, sheet2)
    assert result["status"] == "Ok"
    assert "have the same structure" in result["description"]
    assert not result["errors"]  # Ensure no errors

# pylint: disable=W0621
def test_check_sheet_structure_different_columns(sheet1, sheet_with_different_headers):
    """Test that two sheets with different headers return False and provide the right message."""
    result = check_sheet_structure(sheet1, sheet_with_different_headers)
    assert result["status"] == "Error"
    assert "The following discrepancies were found in the sheet structure:" in result["description"]
    assert "Column 1: Name != First Name" in result["errors"]["Header Mismatch"]

# pylint: disable=W0621
def test_check_sheet_structure_empty_sheet1(empty_sheet, sheet2):
    """Test that an empty sheet returns False and provides the correct message."""
    empty_sheet.title = "Sheet1"
    result = check_sheet_structure(empty_sheet, sheet2)
    assert result["status"] == "Error"
    assert "Sheet1" in result["errors"]["Empty Sheet"]

def test_check_sheet_structure_empty_sheet2(sheet1, empty_sheet):
    """Test that an empty sheet returns False and provides the correct message."""
    empty_sheet.title = "Sheet2"
    result = check_sheet_structure(sheet1, empty_sheet)
    assert result["status"] == "Error"
    assert "Sheet2" in result["errors"]["Empty Sheet"]

def test_check_sheet_structure_different_size(sheet1, sheet_with_different_headers):
    """Test that two sheets with different sizes (number of rows/columns) return False."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Sheet2'
    sheet['A1'] = 'Name'
    sheet['B1'] = 'Age'
    sheet['C1'] = 'City'  # Extra column
    sheet['A2'] = 'Alice'
    sheet['B2'] = 30
    sheet['C2'] = 'NYC'
    sheet['A3'] = 'Bob'
    sheet['B3'] = 20
    sheet['C3'] = 'LON'
    sheet_with_different_headers = sheet

    result = check_sheet_structure(sheet1, sheet_with_different_headers)
    assert result["status"] == "Error"
    assert "The following discrepancies were found in the sheet structure:" in result["description"]
    assert "'Sheet1' has 3 rows and 2 columns, 'Sheet2' has 3 rows and 3 columns."\
        in result["errors"]["Row/Column Count"]
    assert "'Sheet1' has 3 rows and 2 columns, 'Sheet2' has 3 rows and 3 columns."\
        in result["errors"]["Row/Column Count"]

def test_check_sheet_structure_invalid_input():
    """Test that invalid inputs return the expected error message."""
    with pytest.raises(ValueError):
        check_sheet_structure(None, None)
