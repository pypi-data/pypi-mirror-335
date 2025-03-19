# -*- coding: UTF-8 -*-
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter


class CreateExcel:
    """CreateExcel"""

    def __init__(self, msg):
        # 创建五个表
        self.workbook = Workbook()
        self.workbook.active.title = "升级版本软件包"
        self.workbook.create_sheet("降低版本软件包")
        self.workbook.create_sheet("同版本软件包")
        self.workbook.create_sheet("删除软件包")
        self.workbook.create_sheet("新增软件包")
        self.workbook.create_sheet("文件目录比对")
        self.workbook.create_sheet("文件内容比对")

        # 填充
        self.blue_fill = PatternFill('solid', fgColor='4F81BD')

        # 顶部样式
        self.header_title = NamedStyle(name="header_title")
        self.header_title.font = Font(name=u"宋体", sz=14, bold=True)
        self.header_title.alignment = Alignment(horizontal='center', vertical="center")
        self.header_title.fill = self.blue_fill

        # 正文样式
        self.header_name = NamedStyle(name="header_name")
        self.header_name.font = Font(name=u"宋体", sz=14)
        self.header_name.alignment = Alignment(vertical="center", wrap_text=True)

        # 添加顶部信息
        for i in range(0, 7):
            self.__add_header__(i)
            self.workbook.active.cell(2, 1).value = msg
        self.workbook.active = 0

    def __add_header_up_down__(self, index):
        self.workbook.active = index
        self.workbook.active.merge_cells('A1:D1')
        self.workbook.active.merge_cells('A2:D2')
        self.workbook.active.merge_cells('A3:B3')
        self.workbook.active.merge_cells('C3:D3')
        self.workbook.active.cell(3, 1, value='SRPM').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(3, 3, value='RPM').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(1, 1).value = self.workbook.sheetnames[index]
        self.workbook.active.cell(1, 1).style = self.header_title
        self.workbook.active.cell(2, 1).style = self.header_name
        for i in range(1, 5):
            if (i % 2) == 0:
                self.workbook.active.cell(4, i, 'B').alignment = Alignment(horizontal='center', vertical='center')
            else:
                self.workbook.active.cell(4, i, 'A').alignment = Alignment(horizontal='center', vertical='center')
            self.workbook.active.column_dimensions[get_column_letter(i)].width = 70
            if i < 3:
                self.workbook.active.row_dimensions[i].height = 40

    def __add__header_add_del__(self, index):
        self.workbook.active = index
        self.workbook.active.merge_cells('A1:B1')
        self.workbook.active.merge_cells('A2:B2')
        self.workbook.active.cell(1, 1).value = self.workbook.sheetnames[index]
        self.workbook.active["A1"].style = self.header_title
        self.workbook.active["A2"].style = self.header_name
        for i in range(1, 3):
            self.workbook.active.column_dimensions[get_column_letter(i)].width = 70
            self.workbook.active.row_dimensions[i].height = 40
        self.workbook.active.cell(3, 1, value='SRPM').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(3, 2, value='RPM').alignment = Alignment(horizontal='center', vertical='center')

    def __add_header_list__(self, index):
        self.workbook.active = index
        self.workbook.active.merge_cells('A1:C1')
        self.workbook.active.merge_cells('A2:C2')
        self.workbook.active.cell(1, 1).value = self.workbook.sheetnames[index]
        self.workbook.active["A1"].style = self.header_title
        self.workbook.active["A2"].style = self.header_name
        for i in range(1, 4):
            if i == 1:
                self.workbook.active.column_dimensions[get_column_letter(i)].width = 50
            else:
                self.workbook.active.column_dimensions[get_column_letter(i)].width = 35
            if i != 3:
                self.workbook.active.row_dimensions[i].height = 40
        self.workbook.active.cell(3, 1, value='文件名').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(3, 2, value='A(md5)').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(3, 3, value='B(md5)').alignment = Alignment(horizontal='center', vertical='center')

    def __add_header_content__(self, index):
        self.workbook.active = index
        self.workbook.active.merge_cells('A1:B1')
        self.workbook.active.merge_cells('A2:B2')
        self.workbook.active.cell(1, 1).value = self.workbook.sheetnames[index]
        self.workbook.active["A1"].style = self.header_title
        self.workbook.active["A2"].style = self.header_name
        self.workbook.active.column_dimensions['A'].width = 30
        self.workbook.active.row_dimensions[1].height = 40
        self.workbook.active.column_dimensions['B'].width = 100
        self.workbook.active.row_dimensions[2].height = 40
        self.workbook.active.cell(3, 1, value='文件名').alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.active.cell(3, 2, value='差异内容').alignment = Alignment(horizontal='center', vertical='center')

    def __add_header__(self, index):
        if index < 0 or index > 7:
            return "参数错误"

        if 0 <= index < 2:
            self.__add_header_up_down__(index)
        elif 2 <= index <= 4:
            self.__add__header_add_del__(index)
        elif index == 5:
            self.__add_header_list__(index)
        else:
            self.__add_header_content__(index)

    def add_compare_info(self, index, info_a, info_b, compare_type):
        """
        添加源码包信息
        @param index: 表
        @param info_a:  添加信息a
        @param info_b: 添加信息b
        @param compare_type: 信息类型
        @return:
        """
        if index < 0 or index >= 2:
            return "参数有误"
        self.workbook.active = index
        if compare_type.lower() == "srpm":
            for i in range(0, len(info_a)):
                self.workbook.active.cell(i + 5, 1).value = info_a[i]
            for i in range(0, len(info_b)):
                self.workbook.active.cell(i + 5, 2).value = info_b[i]
        elif compare_type.lower() == 'rpm':
            for i in range(0, len(info_a)):
                self.workbook.active.cell(i + 5, 3).value = info_a[i]
            for i in range(0, len(info_b)):
                self.workbook.active.cell(i + 5, 4).value = info_b[i]

    def add_common(self, index, info, info_type):
        """
        @param index: 表
        @param info: 添加信息
        @param info_type:  添加信息类型
        @return:
        """
        if index <= 1 or index > 4:
            return "参数有误"
        self.workbook.active = index
        if info_type.lower() == 'srpm':
            for i in range(0, len(info)):
                self.workbook.active.cell(i + 4, 1).value = info[i]
        elif info_type.lower() == 'rpm':
            for i in range(0, len(info)):
                self.workbook.active.cell(i + 4, 2).value = info[i]
        self.workbook.active = 0

    def add_files(self, index, files, a_md5, b_md5):
        if index != 5:
            return "参数有误"
        self.workbook.active = index
        for _ in range(0, len(files)):
            self.workbook.active.cell(_ + 4, 1).value = files[_]
            self.workbook.active.cell(_ + 4, 2).value = a_md5[_]
            self.workbook.active.cell(_ + 4, 3).value = b_md5[_]
        self.workbook.active = 0

    def add_content_diff(self, index, files, contents, initrd_diff):
        if index != 6:
            return "参数有误"
        self.workbook.active = index
        for _ in range(0, len(files) + 1):
            if _ < len(files):
                self.workbook.active.cell(_ + 4, 1).value = files[_]
                self.workbook.active.cell(_ + 4, 2).value = contents[_]
            else:
                self.workbook.active.cell(_ + 4, 1).value = initrd_diff['fp']
                self.workbook.active.cell(_ + 4, 2).value = initrd_diff['diff_content']
            self.workbook.active.cell(_ + 4, 2).alignment = Alignment(wrap_text=True)
        self.workbook.active = 0

    def save(self, save_path):
        """
        保存表
        @param save_path: 保存路径
        @return:
        """
        if os.path.exists(save_path):
            print(save_path + " 此文件会被覆盖。------------")
        self.workbook.save(save_path)
