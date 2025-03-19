"""Support for writing tabular data to a file or data stream, in CSV or XLSX."""

import csv
from io import BytesIO
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from ..sys_utils.log_tools import setup_logger

logger = setup_logger(__name__)
# pylint: disable=logging-fstring-interpolation


def write_data_table(
    data: list[dict[str, str]],
    file_path: str | None,
    column_widths: dict[str, int] | None = None,
) -> BytesIO | None:
    """
    Write dictionaries to a file (CSV or XLSX, based on the file extension).

    Args:
        data: List of dictionaries containing the tabular data.
        file_path: Path to the output file. If None, defaults to an XLSX file in memory.
        column_widths: Optional dictionary specifying the width of each column.

    Returns:
        A BytesIO object containing the XLSX data if file_path is None and the
        format is XLSX, otherwise None.
    """
    if file_path is None:
        # Default to .xlsx if file_path is None
        return write_dicts_to_xlsx(data, None, column_widths=column_widths)

    _, file_extension = os.path.splitext(file_path)

    if file_extension.lower() == ".csv":
        write_dicts_to_csv(data, file_path)
        return None
    if file_extension.lower() == ".xlsx":
        return write_dicts_to_xlsx(data, file_path, column_widths=column_widths)
    raise ValueError("Unsupported file type")


def write_dicts_to_csv(data: list[dict[str, str]], file_path: str) -> None:
    """
    Write a list of dictionaries to a CSV file.

    Args:
        data: List of dictionaries containing the tabular data.
        file_path: Path to the output CSV file.

    Raises:
        ValueError: If file_path is empty.
    """
    if not file_path:
        raise ValueError("Write of CSV data to memory buffer is not supported.")
    with open(file_path, mode="w", encoding="utf-8", newline="") as file:
        if data:
            fieldnames = data[0].keys()
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)
        else:
            logger.warning("No data to write.")


def write_dicts_to_xlsx(
    data: list[dict[str, str]],
    file_path: str | None = None,
    column_widths: dict[str, int] | None = None,
) -> BytesIO | None:
    """
    Write a list of dictionaries to an XLSX file.

    Args:
        data: List of dictionaries containing the tabular data.
        file_path: Path to the output XLSX file. If None, writes to a BytesIO object.
        column_widths: Optional dictionary specifying the column order and widths.

    Returns:
        A BytesIO object containing the XLSX data if file_path is None, otherwise None.
    """
    workbook = Workbook()
    sheet = workbook.active
    if data:
        headers = list(column_widths.keys()) if column_widths else list(data[0].keys())
        sheet.append(headers)
        header_font = Font(bold=True)
        sheet.row_dimensions[1].height = 30
        header_alignment = Alignment(wrap_text=True, vertical="center")

        for i, header in enumerate(headers, start=1):
            column_letter = get_column_letter(i)
            cell = sheet[f"{column_letter}1"]
            cell.font = header_font
            cell.alignment = header_alignment

        for i, header in enumerate(headers, start=1):
            column_letter = get_column_letter(i)
            if column_widths and header in column_widths:
                sheet.column_dimensions[column_letter].width = column_widths[header]

        for row in data:
            # Write the values in the order of the headers
            row_values = [row.get(key, "") for key in headers]
            sheet.append(row_values)
    else:
        print("No data to write.")

    if file_path:
        workbook.save(file_path)
        return None
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    return excel_file
