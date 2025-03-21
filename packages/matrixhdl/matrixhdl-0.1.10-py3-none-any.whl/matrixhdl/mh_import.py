##/////////////////////////////////////////////////////////////////////////////////////
##//
##//Copyright 2025  Li Xinbing
##//
##//Licensed under the Apache License, Version 2.0 (the "License");
##//you may not use this file except in compliance with the License.
##//You may obtain a copy of the License at
##//
##//    http://www.apache.org/licenses/LICENSE-2.0
##//
##//Unless required by applicable law or agreed to in writing, software
##//distributed under the License is distributed on an "AS IS" BASIS,
##//WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
##//See the License for the specific language governing permissions and
##//limitations under the License.
##//
##/////////////////////////////////////////////////////////////////////////////////////

###################################################################################
##import
###################################################################################
import openpyxl    ##openpyxl.load_workbook()
import os          ##os.path.exists()/os.makedirs()/os.getcwd()
import shutil      ##shutil.rmtree()
import re
###################################################################################
##class:SysArgv
###################################################################################
class SysArgv:
    ''' To analyze sys.argv and generate parameters. '''
    from sys import argv
    xlsx_file = 'rtl.xlsx' 
    rtl_files = [] 
    top       = ''
    level     = 0
    
    def __init__(self):
        del self.argv[0]
        given_para_title = ['-o','-f','-v','-top','-level','-full']
        given_para_value = [self.get_extract_name(i) for i in given_para_title]
        oname = given_para_value[0]
        fname = given_para_value[1]
        vname = given_para_value[2]
        top   = given_para_value[3]
        level = given_para_value[4]
        full  = given_para_value[5]
        files = [i for i in self.argv if i not in given_para_title+given_para_value]
        if vname:
            for root,folder_names,file_names in os.walk(vname):
                for file_name in file_names:
                    if file_name.endswith('.v'):
                        files.append(root+'\\'+file_name)
        if fname:
            with open(fname) as f:
                other = f.read().strip().split()
            for i in other:
                if not any(i.startswith(each) for each in ['+','-','//']):
                    files.append(i)
        self.xlsx_file = oname if oname else self.xlsx_file
        self.rtl_files = files
        self.top   = top if top else ''
        self.level = level if level else 0
        self.full  = full if full else False
        if not files:
            print("---No input .v files,exit...")
            exit()
        
    def get_extract_name(self,char):
        if char in self.argv:
            if self.argv.index(char)<((len(self.argv))-1) and not self.argv[self.argv.index(char)+1].startswith('-'):
                value = self.argv[self.argv.index(char)+1]
            else:
                value = True
        else:
            value = False
        return value
            
    def do_make_path(self,path):
        if os.path.exists(path):
            shutil.rmtree(path)
        os.makedirs(path)
        
    def do_check_path(self,path):
        if not os.path.exists(path):
            os.makedirs(path)    

###################################################################################
##class:AllVfiles
###################################################################################
class AllVfiles:
    '''The class of all .v files'''
    xlsx_file   = ''
    top         = ''
    level       = 0
    full        = False
    text_raw    = {}
    text_clean  = {}
    m           = {}
    group_member= []
    group_parent= []
    x           = {}
    branch      = []
    branch_left = []
    def __init__(self,argv):
        self.xlsx_file  = argv.xlsx_file
        self.top        = argv.top
        self.level      = int(argv.level)
        self.full       = argv.full
        print('----Begin loading .v files...')
        self.text_raw   = self.get_raw_text(argv.rtl_files)  
        print('----Cleaning comments...')        
        self.text_clean = self.get_clean_text(self.text_raw)
        print('----Parsing .v files and getting its structure...')
        self.m          = self.get_initial_m_dict(self.text_clean)
        print('------parameters and ports...')
        self.do_get_paralist_and_portlist()
        print('------contents...')
        self.do_divide_bodyfull()
        print('------compiling ports...')
        self.do_get_port_type_width()
        print('------parameters,regs,wires...')
        self.do_get_para_reg_wire()
        print('------submodules...')
        self.do_get_submodule()
        print('----Preparing matrixes of .v files...')
        self.group_member,self.group_parent = self.do_get_module_tree()
        self.do_initial_matrix()
        for cnt,parent in enumerate(self.group_parent):
            print(parent,'--------',self.group_member[cnt])
        print('len of all modules are ',len(self.m.keys()))
        print('len of all group_member ',len([j for i in self.group_member for j in i]))
        print('len of groups ',len(self.group_member))                    
        self.do_write_xlsx()
    
    def do_initial_matrix(self):
        for member in self.group_member:
            for module in member:
                self.do_initial_module(module)
        
    def do_initial_module(self,n):
        self.x[n] = {}
        module = self.m[n]
        self.x[n]['colnum'] = 2+len(module['sub'])
        self.x[n]['array'] = []
        ##1 row
        self.x[n]['array'].append([n,'']+[module['sub'][i]['ref']+'  '+i for i in module['order']])
        ##2 row: paralist
        self.x[n]['array'].append([module['paralist'],'']+[module['sub'][i]['paralist'] for i in module['order']])
        ##3,row: file
        self.x[n]['array'].append(['','']+['' for i in range(len(module['sub']))])
        self.x[n]['array'].append(['' for i in range(self.x[n]['colnum'])])
        ##4 row -- parameter
        if module['paradef']:
            self.x[n]['array'].append(['//parameter','//']+['//' for i in range(len(module['sub']))])
            for para in module['paradef']:
                self.x[n]['array'].append([para,'']+['' for i in range(len(module['sub']))])
        ##5 row --- connection
        endpoint = []
        if module['sub']:
            self.x[n]['array'].append(['//connection start','//']+['//' for i in range(len(module['sub']))])
            self.x[n]['array'].append(['' for i in range(self.x[n]['colnum'])])
            for sub_cnt,sub_one in enumerate(module['order']):
                sub    = module['sub'][sub_one]
                insnet = sub['insnet']
                iox = {}
                for i,one in enumerate(sub['refdef']):
                    iox[one] = (10000 if '#i' in one else 20000 if '#o' in one else 30000 if '#x' in one else 40000)+i
                refdef = [i[0] for i in sorted(iox.items(),key=lambda d:d[1],reverse=False)]
                refdef_count = [i[1] for i in sorted(iox.items(),key=lambda d:d[1],reverse=False)]                
                signal_name   = module['portname']+module['wirename']+module['regname']
                signal_def    = module['portdef']+module['wiredef']+module['regdef']
                signal_assign = module['portassign']+module['wireassign']+['' for i in range(len(module['regname']))]
                for net_cnt,net_one in enumerate(refdef):
                    net_cnt = refdef_count[net_cnt]%10000
                    net_one = insnet[net_cnt]
                    if not net_one in endpoint and net_one in signal_name:
                        net_one_in_submodule = [module['sub'][i]['insnet'].count(net_one) for i in module['order'][sub_cnt:]]
                        array = [['' for i in range(self.x[n]['colnum'])] for j in range(max(net_one_in_submodule))]
                        array[0][0] = signal_def[signal_name.index(net_one)]
                        array[0][1] = signal_assign[signal_name.index(net_one)]
                        for num,one in enumerate(module['order'][sub_cnt:]):
                            nets_order = [cnt for cnt,net in enumerate(module['sub'][one]['insnet']) if net==net_one] 
                            for cnt,order in enumerate(nets_order):
                                if cnt==0:
                                    result = module['sub'][one]['refdef'][order] if module['sub'][one]['inspin'][order]==module['sub'][one]['insnet'][order] else module['sub'][one]['refpin'][order]+'  '+module['sub'][one]['inspin'][order].replace(net_one,'-')
                                else:
                                    result = module['sub'][one]['refpin'][order]+'  '+module['sub'][one]['inspin'][order]
                                array[cnt][2+sub_cnt+num] = result
                        self.x[n]['array'].extend(array)
                        endpoint.append(net_one)
                    elif not net_one:
                        array = ['' for i in range(self.x[n]['colnum'])]
                        array[2+sub_cnt] = sub['refpin'][net_cnt]+'  '+sub['inspin'][net_cnt]
                        self.x[n]['array'].append(array)                          
            self.x[n]['array'].append(['' for i in range(self.x[n]['colnum'])])
            self.x[n]['array'].append(['//connection over','//']+['//' for i in range(len(module['sub']))])            
        ##6 row body
        rest_port       = [module['portdef'][cnt] for cnt,i in enumerate(module['portname']) if i not in endpoint]
        rest_portassign = [module['portassign'][cnt] for cnt,i in enumerate(module['portname']) if i not in endpoint]
        rest_wire = [module['wiredef'][cnt] for cnt,i in enumerate(module['wirename']) if i not in endpoint]
        rest_wireassign = [module['wireassign'][cnt] for cnt,i in enumerate(module['wirename']) if i not in endpoint]
        rest_reg  = [module['regdef'][cnt] for cnt,i in enumerate(module['regname']) if i not in endpoint]   
        rest_len = len(rest_port+rest_wire+rest_reg)
        body_split = (module['bodyone']+module['bodyfunc']).strip().splitlines()
        body_len = len(body_split)
        length = rest_len+body_len
        if length:
            self.x[n]['array'].append(['' for i in range(self.x[n]['colnum'])])        
            self.x[n]['array'].append(['//body start','//']+['//' for i in range(len(module['sub']))])
            self.x[n]['array'].append(['' for i in range(self.x[n]['colnum'])])             
            self.x[n]['array'].append(['' for i in range(self.x[n]['colnum'])])
            for i in range(len(rest_port)):
                self.x[n]['array'].append([rest_port[i],rest_portassign[i]]+['' for i in range(len(module['sub']))]) 
            for i in range(len(rest_wire)):
                self.x[n]['array'].append([rest_wire[i],rest_wireassign[i]]+['' for i in range(len(module['sub']))])
            for i in range(len(rest_reg)):
                self.x[n]['array'].append([rest_reg[i],'']+['' for i in range(len(module['sub']))]) 
            if self.full or (body_len<1000):
                for i in range(body_len):
                    self.x[n]['array'].append(['',body_split[i]]+['' for i in range(len(module['sub']))])   
            else:
                print('-----',n,' module  has more than 1000 lines(%s) left, ignoring...'%body_len) 
            self.x[n]['array'].append(['' for i in range(self.x[n]['colnum'])])        
            self.x[n]['array'].append(['//body over','//']+['//' for i in range(len(module['sub']))])
            self.x[n]['array'].append(['' for i in range(self.x[n]['colnum'])])         
        self.x[n]['array'].append(['' for i in range(self.x[n]['colnum'])])
        self.x[n]['array'].append(['' for i in range(self.x[n]['colnum'])])               
        self.x[n]['array'].append(['//'+module['file'],'']+['' for i in range(len(module['sub']))])
        self.x[n]['array'].append(['' for i in range(self.x[n]['colnum'])])           
        self.x[n]['rownum'] = len(self.x[n]['array'])    
           
    def get_raw_text(self,files):
        result = {}
        print('---Files are...')
        for file in files:
            print('---',file)
            #with open(file) as f:
            with open(file,encoding='utf-8') as f:
                result[file] = f.read()
        return result
    
    def get_clean_text(self,text):
        for name,body in text.items():
            nearest_char,nearest_get,nearest_pos = self.get_nearest_chars(body,['/*','//','`i','`e'])
            while nearest_get:
                if nearest_char=='/*':
                    body = self.get_remove_slash_star(body)
                elif nearest_char=='`i':
                    body = self.get_remove_define_i(body)
                elif nearest_char=='`e':
                    body = self.get_remove_define_e(body)
                else:
                    body = self.get_remove_two_slashes(body)
                nearest_char,nearest_get,nearest_pos = self.get_nearest_chars(body,['/*','//','`i','`e'])
            text[name] = body
        return text   
    
    def get_initial_m_dict(self,text):
        result = {}
        for file,modules in text.items():
            for module in modules.split('endmodule'):
                if module.find('module ')!=-1:
                    module = module[module.index('module ')+len('module '):]
                    name   = re.match(r'[a-zA-Z]\w*',module.strip()).group()
                    result[name] = {}
                    result[name]['file'] = file
                    result[name]['text'] = module[len(name):].strip()
        return result
        
    def do_get_paralist_and_portlist(self):
        for top_name,module in self.m.items():
            print(top_name)
            text     = module['text'].strip(' ;')
            paralist = ''
            if text.startswith('#'):
                text,paralist = self.get_one_part(text)
            text,portlist = self.get_one_part(text)
            self.m[top_name]['paralist'] = '#('+' '.join(paralist.strip().split())+')' if paralist else ''
            self.m[top_name]['portlist'] = portlist
            self.m[top_name]['portname'] = [i.split()[-1] for i in portlist.split(',')] if portlist.strip() else ''
            self.m[top_name]['bodyfull'] = text.strip().lstrip(';')
    
    def do_divide_bodyfull(self):
        for top_name,module in self.m.items():
            body = self.m[top_name]['bodyfull']
            func = ''
            for keyword in ['function','task','generate']:
                body,block = self.get_remove_block(body,keyword)
                func += block
            self.m[top_name]['bodyfunc'] = func
            self.m[top_name]['bodyrest'] = body
    
    def do_get_port_type_width(self):
        for top_name,module in self.m.items():
            portlist = module['portlist']
            portname = module['portname']
            bodyrest = module['bodyrest']
            portnum  = len(portname)
            self.m[top_name]['porttype']  = ['' for i in range(portnum)]
            self.m[top_name]['portwidth'] = ['' for i in range(portnum)] 
            self.m[top_name]['portdef']   = ['' for i in range(portnum)]    
            if re.search(r'\b(input|output|inout)\b',portlist):
                groups = portlist.split(',')
            else:
                result = re.findall(r'\binput\b[^;]*;|\boutput\b[^;]*;|\binout\b[^;]*;',bodyrest)
                for i in result:
                    bodyrest = bodyrest.replace(i,'')
                groups = [i for group in result for i in group.strip().split(',')] 
            for one in groups:
                one = one.rstrip(';').strip()
                for port_type in ['input','output','inout']:
                    if re.match( fr'{port_type}(?!\w)',one):
                        type  = '#i' if port_type=='input' else '#o' if port_type=='output' else '#x'
                        width = one[one.index('['):one.index(']')+1] if '[' in one else ''  
                signal = one.split()[-1]
                signal = signal[signal.index(']')+1:] if ']' in signal else signal
                self.m[top_name]['porttype'][portname.index(signal)]  = type
                self.m[top_name]['portwidth'][portname.index(signal)] = width
                self.m[top_name]['portdef'][portname.index(signal)]   = (signal+' '+type+' '+width).strip()
            self.m[top_name]['bodymain'] = bodyrest.strip()
    
    def do_get_para_reg_wire(self):
        for top_name,module in self.m.items():
            bodymain = module['bodymain'] 
            self.m[top_name]['paraname']  = []
            self.m[top_name]['parawidth'] = []
            self.m[top_name]['paraval']   = []
            self.m[top_name]['paradef']   = []
            width = ''
            result = re.findall(r'\bparameter\b\s*[^;]*;|\blocalparam\b\s+[^;]*;',bodymain)
            for group in result:
                bodymain = bodymain.replace(group,'')
                for one in group.split(','):
                    if '=' in one:
                        one = one.rstrip(';').strip()
                        width = (one[one.index('['):one.index(']')+1] if '[' in one else '') if any(one.startswith(i) for i in ['parameter ','localparam']) else width
                        signal= one.split('=')[0].split()[-1]
                        val   = ' '.join(''.join(one.split('=')[1:]).split())
                        self.m[top_name]['paraname'].append(signal)
                        self.m[top_name]['parawidth'].append(width)
                        self.m[top_name]['paraval'].append(val)
                        self.m[top_name]['paradef'].append((signal+' #p'+' '+width+' '+val).strip())
                    else:
                        self.m[top_name]['paraval'][-1] += ','+one.rstrip(';')
            self.m[top_name]['regname']  = []
            self.m[top_name]['regwidth'] = []
            self.m[top_name]['regdef']   = []
            result = re.findall(r'\breg\b\s+[^=;]*;',bodymain)
            for group in result:
                bodymain = bodymain.replace(group,'')
                for one in group.split(','):
                    one    = one.rstrip(';').strip()
                    width  = (one[one.index('['):one.index(']')+1] if '[' in one else '') if one.startswith('reg') else width
                    signal = one.split()[-1]
                    self.m[top_name]['regname'].append(signal)
                    self.m[top_name]['regwidth'].append(width)
                    self.m[top_name]['regdef'].append((signal+' #r'+' '+width).strip())    
            self.m[top_name]['wirename']  = []
            self.m[top_name]['wirewidth'] = []
            self.m[top_name]['wiredef']   = []
            result = re.findall(r'\bwire\b\s+[^=;]*;',bodymain)
            for group in result:
                bodymain = bodymain.replace(group,'')
                for one in group.split(','):
                    one    = one.rstrip(';').strip()
                    width  = (one[one.index('['):one.index(']')+1] if '[' in one else '') if one.startswith('wire') else width
                    signal = one.split()[-1]
                    if signal not in self.m[top_name]['portname']:
                        self.m[top_name]['wirename'].append(signal)
                        self.m[top_name]['wirewidth'].append(width)
                        self.m[top_name]['wiredef'].append((signal+' #w'+' '+width).strip())                                    
            self.m[top_name]['wireassign'] = ['' for i in range(len(self.m[top_name]['wirename']))]
            self.m[top_name]['portassign'] = ['' for i in range(len(self.m[top_name]['portname']))]
            result = re.findall(r'\bassign\b\s+[a-zA-Z]\w*[^;]*;',bodymain)
            name   = [re.match(r'\w*',i[len('assign '):].strip()).group() for i in result]
            count  = [name.count(i)==1 for i in name]
            name   = [name[i] for i in range(len(result)) if count[i]]              
            result = [result[i] for i in range(len(result)) if count[i]]   
            for cnt,one in enumerate(result):
                bodymain = bodymain.replace(one,'')
                one      = one.strip()
                signal   = name[cnt]
                if signal in self.m[top_name]['wirename']:
                    self.m[top_name]['wireassign'][self.m[top_name]['wirename'].index(signal)] = '# '+one[one.index('='):]
                elif signal in self.m[top_name]['portname']:
                    self.m[top_name]['portassign'][self.m[top_name]['portname'].index(signal)] = '# '+one[one.index('='):]   
                else:
                    self.m[top_name]['wirename'].append(signal)
                    self.m[top_name]['wirewidth'].append('')
                    self.m[top_name]['wiredef'].append(signal+' #w')
                    self.m[top_name]['wireassign'].append('# '+one[one.index('='):])
            self.m[top_name]['bodyinst'] = bodymain.strip()
    
    def do_get_submodule(self):
        print('------Submodule progress...')
        for top_name,module in self.m.items():
            print(top_name)
            bodyinst = module['bodyinst']
            self.m[top_name]['sub'] = {}
            keys = list(self.m.keys())[:]
            del keys[keys.index(top_name)]
            for key in keys:
                result = re.search(r'\b%s\b\s+'%key,bodyinst)
                while result:
                    head = bodyinst[:result.start()]
                    text = bodyinst[result.start()+len(key+' '):].strip()
                    paralist = ''
                    if text.startswith('#'):
                        text,paralist = self.get_one_part(text)
                    ins_name = re.match(r'\w+',text.strip()).group()
                    self.m[top_name]['sub'][ins_name] = {}
                    text,portlist = self.get_one_part(text)
                    bodyinst = head + text.strip().lstrip(';').strip()
                    self.m[top_name]['sub'][ins_name]['ref']      = key
                    self.m[top_name]['sub'][ins_name]['paralist'] = '#('+' '.join(paralist.strip().split())+')' if paralist else ''
                    self.m[top_name]['sub'][ins_name]['portlist'] = portlist
                    refpin,inspin = self.get_analyze_portlist(portlist)
                    self.m[top_name]['sub'][ins_name]['refpin'] = refpin
                    self.m[top_name]['sub'][ins_name]['refdef'] = [self.m[key]['portdef'][self.m[key]['portname'].index(pin)]  for pin in refpin]
                    self.m[top_name]['sub'][ins_name]['inspin'] = inspin
                    self.m[top_name]['sub'][ins_name]['insnet'] = [re.match(r'[a-zA-Z]\w*',one).group() if re.match(r'[a-zA-Z]\w*',one) else '' for one in inspin]
                    result = re.search(r'\b%s\b\s+'%key,bodyinst)
            order = {}
            for key in self.m[top_name]['sub'].keys():
                order[key] = len(self.m[top_name]['sub'][key]['refpin'])
                ##order[key] = re.search(r'\b%s\b'%key,module['bodyinst']).start()
            self.m[top_name]['order'] = [i[0] for i in sorted(order.items(),key=lambda d:d[1],reverse=False)]
            self.m[top_name]['bodyone'] = '\n'.join([one for one in bodyinst.splitlines() if one.strip()])
    
    def do_get_module_tree(self):
        module    = list(self.m.keys())
        submodule = [[self.m[one]['sub'][i]['ref'] for i in self.m[one]['order']] for one in module]
        sub_all   = [j for i in submodule for j in i]
        self.branch    = [i for i in module if i not in sub_all]
        group_member  = [[i] for i in module if (not self.top and i not in sub_all) or (self.top and (i==self.top))]
        self.branch_left = [i for i in self.branch if i not in [j for i in group_member for j in i]]
        group_parent  = ['TOP'+str(i) for i in range(len(group_member))]      
        group_level   = [0 for i in range(len(group_member))]
        current   = 0
        while (current<len(group_member)) and ((not self.level) or (self.level and (group_level[current]<self.level))):
            for parent in group_member[current]:
                sub = [i for i in submodule[module.index(parent)] if i not in [m for n in group_member for m in n]]
                member = []
                for one in sub:
                    if one not in member and (self.full or self.m[one]['sub']):
                        member.append(one)
                if any(member):
                    group_member.append(member)
                    group_parent.append(parent)
                    group_level.append(group_level[current]+1)
            current += 1
        return group_member,group_parent    
    
    def get_analyze_portlist(self,portlist):
        refpin = []
        inspin = []
        portlist = portlist.strip()
        while portlist.startswith('.'):
            ref = re.match(r'\w*',portlist[1:].strip()).group()
            portlist,ins = self.get_one_part(portlist) if portlist else ''
            refpin.append(ref)
            inspin.append(ins.strip())
            portlist = portlist.strip('').lstrip(',').strip()
        return refpin,inspin
    
    def get_nearest_chars(self,string,chars):
        pos = [self.get_pos(string,char) for char in chars]
        idx = pos.index(min(pos))
        return chars[idx],pos[idx]!=len(string),pos[idx]
    
    def get_remove_slash_star(self,string):
        start = string.find('/*')
        end = start + string[start:].find('*/')
        string = string[:start]+string[end+2:]
        return string
     
    def get_remove_two_slashes(self,string):
        start = string.find('//')
        end   = start + string[start:].find('\n')
        string = string[:start]+string[end:]
        return string  
      
    def get_remove_define_i(self,string):
        start = string.find('`i')
        end   = start + string[start:].find('\n')
        string = string[:start]+string[end:]
        return string        
       
    def get_remove_define_e(self,string):
        start = string.find('`e')
        end   = start + string[start:].find('\n')
        string = string[:start]+string[end:]
        return string          
    
    def get_one_part(self,text):
        result  = ''
        if text:
            text    = self.rem_header(text,'(')
            cnt     = 1
            while cnt:
               open  = self.get_pos(text,'(')
               close = self.get_pos(text,')')
               if open<close:
                   result += text[:open+1]
                   text    = text[open+1:]
                   cnt    += 1
               else:
                   result += text[:close+1]
                   text    = text[close+1:]
                   cnt    -= 1
        return text,result.rstrip(')')         
    
    def get_remove_block(self,text,block):
        generate = ''
        while re.search(r'\b%s\b\s+'%block,text) and re.search(r'\bend%s\b'%block,text):
            start     = re.search(r'\b%s\b\s+'%block,text).start()
            end       = re.search(r'\bend%s\b'%block,text).end()+1
            generate += text[start:end]+'\n'
            text      = text[:start]+text[end:]
        return text,generate  
    
    rem_header = lambda self,text,header: text[text.index(header)+len(header):]
    get_pos    = lambda self,text,char  : len(text) if text.find(char)==-1 else text.find(char)
    
    ####################################################################################################################################
    def do_set_conditional_format(self,ws):
        from openpyxl.styles import PatternFill,Font,Border,Side
        from openpyxl.styles.differential import DifferentialStyle        
        from openpyxl.formatting.rule import Rule
        #border = Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        font = Font(color='000000')
        field = ws.cell(1,1).coordinate+':'+ws.cell(ws.max_row,ws.max_column).coordinate
        #for i in range(1,ws.max_row+1):
        #    for j in range(1,ws.max_column+1):
        #        ws.cell(i,j).border = border
        fill_color = {'//':'F8CBAD','#x':'FFE699','#o':'B4C6E7','#i':'C6E0B4'}
        for char,color in fill_color.items():
            dxf  = DifferentialStyle(font=font ,fill=PatternFill(bgColor=color))#,border=border)
            rule = Rule(type='containsText', operator='containsText', formula=['NOT(ISERROR(SEARCH("%s",A1)))'%(char)], text=char, dxf=dxf)
            ws.conditional_formatting.add(field,rule)
        
    def do_set_sheet_width(self,ws):
        for i in range(1,ws.max_row+1):
            ws.row_dimensions[i].height = 15
        for i in range(1,ws.max_column+1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 25 if i!=2 else 40
    
    def do_set_cell_value(self,ws,row,col,val):
        ws.cell(column=col+1,row=row+1).value = val
        ws.cell(column=col+1,row=row+1).font  = openpyxl.styles.Font(name='Verdana',size=8)
        ws.cell(column=col+1,row=row+1).alignment = openpyxl.styles.Alignment(vertical='center')
    
    def do_fill_module_sheet(self,ws,module):
        x = self.x[module]
        maxcol = x['colnum']
        maxrow = x['rownum']
        for row in range(maxrow):
            for col in range(maxcol):
                self.do_set_cell_value(ws,row,col,x['array'][row][col])  
        self.do_set_sheet_width(ws) 
        self.do_set_conditional_format(ws)        
    
    def do_write_xlsx(self):  
        wb = openpyxl.Workbook()
        del wb['Sheet']
        for cnt,parent in enumerate(self.group_parent):
            ws = wb.create_sheet('#'+'Sub of '+parent)
            ws.sheet_properties.tabColor = "C00000"        
            for module in self.group_member[cnt]:
                ws = wb.create_sheet(module)
                self.do_fill_module_sheet(ws,module)
                ws.freeze_panes = 'B2' 
        ws = wb.create_sheet('@module')
        leaf = [one for one,module in self.m.items() if (one not in [j for i in self.group_member for j in i]) and (one not in self.branch_left)]
        for col,one in enumerate(leaf):
            self.do_set_cell_value(ws,0,col,one)
            row = 0
            for row,port in enumerate(self.m[one]['portdef']): 
                self.do_set_cell_value(ws,row+2,col,port)
            self.do_set_cell_value(ws,row+3,col,'')
            self.do_set_cell_value(ws,row+4,col,'//'+self.m[one]['file'])
        ws.freeze_panes = 'A2'
        wb.save(self.xlsx_file)
        wb.close()

###################################################################################
##Main
###################################################################################
#if __name__ == '__main__': 
def main():
    one_file = AllVfiles(SysArgv())