# This software is dual-licensed under the GNU General Public License (GPL) 
# and a commercial license.
#
# You may use this software under the terms of the GNU GPL v3 (or, at your option,
# any later version) as published by the Free Software Foundation. See 
# <https://www.gnu.org/licenses/> for details.
#
# If you require a proprietary/commercial license for this software, please 
# contact us at jimuflow@gmail.com for more information.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
#
# Copyright (C) 2024-2025  Weng Jing

import openpyxl
import xlrd

from jimuflow.datatypes import Table
from jimuflow.definition import FlowNode
from jimuflow.locales.i18n import gettext
from jimuflow.runtime.execution_engine import PrimitiveComponent, ControlFlow


class ImportExcelAsTableComponent(PrimitiveComponent):

    async def execute(self) -> ControlFlow:
        file_path = self.read_input('filePath')
        sheet_select_type = self.read_input('sheetSelectType')
        if file_path.lower().endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file_path)
            if sheet_select_type == 'by_index':
                sheet_index = int(self.read_input('sheetIndex'))
                sheet = workbook.worksheets[sheet_index - 1]
            else:
                sheet_name = self.read_input('sheetName')
                sheet = workbook[sheet_name]
            use_first_row_as_header = self.read_input('useFirstRowAsHeader')
            if use_first_row_as_header:
                table = Table([cell.value for cell in sheet[sheet.min_row]])
            else:
                table = Table([f'column{i}' for i in range(sheet.min_column, sheet.max_column + 1)])
            start_row = sheet.min_row + 1 if use_first_row_as_header else sheet.min_row
            for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=sheet.min_column,
                                       max_col=sheet.max_column):
                table.rows.append([cell.value for cell in row])
        else:
            workbook = xlrd.open_workbook(file_path)
            if sheet_select_type == 'by_index':
                sheet_index = int(self.read_input('sheetIndex'))
                sheet = workbook.sheet_by_index(sheet_index - 1)
            else:
                sheet_name = self.read_input('sheetName')
                sheet = workbook.sheet_by_name(sheet_name)
            use_first_row_as_header = self.read_input('useFirstRowAsHeader')
            if use_first_row_as_header:
                table = Table([sheet.cell_value(0, i) for i in range(sheet.ncols)])
            else:
                table = Table([f'column{i + 1}' for i in range(sheet.ncols)])
            start_row = 1 if use_first_row_as_header else 0
            for row in range(start_row, sheet.nrows):
                table.rows.append([sheet.cell_value(row, i) for i in range(sheet.ncols)])

        await self.write_output('table', table)
        return ControlFlow.NEXT

    @classmethod
    def display_description(cls, flow_node: FlowNode):
        sheet_select_type = flow_node.input('sheetSelectType')
        if sheet_select_type == 'by_index':
            return gettext(
                'Import excel file ##{filePath}## as table ##{table}##, select sheet by index ##{sheetIndex}##').format(
                filePath=flow_node.input('filePath'), table=flow_node.output('table'),
                sheetIndex=flow_node.input('sheetIndex'))
        else:
            return gettext(
                'Import excel file ##{filePath}## as table ##{table}##, select sheet by name ##{sheetName}##').format(
                filePath=flow_node.input('filePath'), table=flow_node.output('table'),
                sheetName=flow_node.input('sheetName'))
