import logging
import re
from io import BytesIO
from typing import List
from urllib.parse import quote

import openpyxl
from fastapi import UploadFile
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from starlette.responses import StreamingResponse

from fastapi_assistant.core import ExceptionFactory
from fastapi_assistant.libs import ExcelTools


def process_export(headers: List[str]):
    def wrapper(func):
        def inner_wrapper(*args, **kwargs):
            nonlocal headers
            output = BytesIO()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            logging.info(f'导出的headers:{headers}')
            sheet.append(headers)

            file_name = func(sheet)

            red = InlineFont(color='FF0000')
            for i in range(len(headers)):
                key = sheet_key(i)
                value = sheet[f'{key}1'].value
                rich_string1 = CellRichText([TextBlock(red, i) if i == '*' else i for i in value])

                sheet[f'{key}1'] = rich_string1

            width = 3  # 手动加宽的数值
            # 单元格列宽处理
            dims = {}
            for row in sheet.rows:
                for cell in row:
                    if cell.value:
                        cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                        dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
                for col, value in dims.items():
                    sheet.column_dimensions[get_column_letter(col)].width = value + width

            workbook.save(output)
            output.seek(0)
            file_headers = {"content-type": "application/vnd.ms-excel",
                            "content-disposition": 'attachment;filename={}'.format(quote(file_name))}
            return StreamingResponse(output, media_type='xls/xlsx', headers=file_headers)

        return inner_wrapper

    return wrapper


def sheet_key(num: int):
    a, b = divmod(num, 26)
    c = chr(a + 65 - 1) if a > 0 else ''
    return f'{c}{chr(b + 65)}'


async def public_import(db: Session, file: UploadFile, header_func, model, process_item_func, desc: str,
                        delete: bool = True):
    """
    导入数据
    :param db:
    :param file: 文件
    :param header_func: 表头方法
    :param model: 表model
    :param process_item_func: 处理item的方法
    :param desc: 表描述
    :param delete: 是否删除以前的
    :return:
    """
    try:
        headers, columns = header_func()
        tool = ExcelTools(dict(zip(columns, headers)))
        datas = []
        for item in tool.file_to_dict(file):
            data = process_item_func(item)
            if data:
                datas.append(model(**data))
    except ExceptionFactory as e:
        raise e
    except Exception as e:
        logging.error(f'{desc}导入失败：{e}')
        raise ExceptionFactory(detail='文件格式错误，解析内容失败')
    if not datas:
        return
    try:
        with db.begin_nested():
            db.add_all(datas)
            db.flush()
            if delete:
                db.query(model).filter(model.id < datas[0].id).delete()
                db.flush()
    except Exception as e:
        logging.error(f'{desc}导入失败：{e}')
        db.rollback()
        raise ExceptionFactory(detail='导入失败')
    else:
        db.commit()
