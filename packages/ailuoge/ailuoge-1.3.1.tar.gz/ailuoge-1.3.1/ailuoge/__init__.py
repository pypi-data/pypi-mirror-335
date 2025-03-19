import hashlib
import json
import math
import openpyxl
import os
import pymssql
import pymysql
import re
import requests
import sqlalchemy
import time
import zipfile

import pandas as pd
import numpy as np

from openpyxl.utils import get_column_letter
from tqdm import tqdm


class GetEcData:

    def __init__(self, user_name, user_token):
        self.wsdl = 'http://openapi-web.eccang.com/openApi/api/unity'
        self.headers = {
            'ContentType': 'application/json'
        }

        self.user_name = user_name
        self.user_token = user_token

    def __concat_params(self, biz_content: dict, interface_method: str):
        """
        处理post data, 生成key
        :param biz_content:  详细调用参数
        :param interface_method:  易仓的信息 getStockOrderList
        :return:
        """
        post_data = {
            "app_key": self.user_name,
            "biz_content": json.dumps(biz_content),
            "charset": "UTF-8",
            "interface_method": interface_method,
            "nonce_str": "113456",
            "service_id": "E7HPYV",
            "sign_type": "MD5",
            "timestamp": int(time.time() * 1000),
            "version": "v1.0.0"
        }

        # 将字典转化为易仓需要的加密形式
        post_data_str = ''
        for one_key, one_value in zip(post_data.keys(), post_data.values()):
            if type(one_value) == dict:
                one_value = json.dumps(one_value).replace(': ', ':')
            post_data_str += one_key
            post_data_str += '='
            post_data_str += str(one_value)
            post_data_str += '&'

        post_data_str = post_data_str[:-1]
        post_data_str += self.user_token

        # 对组合后的信息进行加密md5
        post_data['sign'] = hashlib.md5(bytes(post_data_str, encoding='utf-8')).hexdigest()

        return post_data

    def __get_data(self, biz_content: dict, interface_method: str, key_word: str = 'data'):
        """
        获取单页数据，把传入的参数转成json格式，向api请求，提取response.text里的data数据
        :param biz_content:
        :param interface_method:
        :param key_word: 要返回的关键词
        :return:
        """
        concated_params = self.__concat_params(biz_content, interface_method)

        # 获取response
        res = requests.post(self.wsdl, json=concated_params, headers=self.headers)
        # 把response的text的json格式转换成字典格式
        try:
            page_info = json.loads(res.text)
            # 打印异常信息
            if page_info['message'] not in ['Success', 'ok']:
                print(page_info['message'])  
                print(f'当前请求的biz_content：{biz_content}')
                print(f'当前请求的interface_method：{interface_method}')
        except:
            page_info = {'message': '系统异常'}
            print(res.text)

        # 判断是否超时
        try:
            # 根据传入的键返回值
            page_data = json.loads(page_info.get('biz_content')).get(key_word)
        except:
            print('系统异常，易仓可能超时')
            print(f'当前请求的biz_content：{biz_content}')
            print(f'当前请求的interface_method：{interface_method}')
            print(page_info)
            page_data = ''

        return page_data

    def get_data(self, biz_content: dict, interface_method: str, special_param: str = None):
        """
        https://open.eccang.com/#/documentCenter?docId=1287&catId=0-225-225,0-177
        获取请求的数据
        :param biz_content:
        :param interface_method:
        :param special_param: 特殊参数，传入该参数后不会尝试获取数据的最大行数，而是遍历biz_content中的该参数列表
        :return:
        """
        # 0 参数设置
        # 默认页数
        if not biz_content.get('page_size'):
            biz_content['page_size'] = 20

        list_df = []
        if not special_param:
            # 1 获取最大页数
            record_rows = self.__get_data(biz_content, interface_method, 'total_count')
            if not record_rows:
                record_rows = self.__get_data(biz_content, interface_method, 'total')
            # 向上取整
            max_page = math.ceil(int(record_rows) / biz_content.get('page_size'))

            # 2 按页获取数据
            print('按页获取数据')
            for i in tqdm(range(1, max_page + 1)):
                time.sleep(5)  # 易仓限制了请求频率，只能自己减少了
                # 2.1 调整键值对
                biz_content['page'] = i
                # 2.2 获取对应页数的数据
                pg_data = self.__get_data(biz_content, interface_method)
                if pg_data:
                    list_df.append(pd.DataFrame(pg_data))
        else:
            list_param = biz_content[special_param]  # 参数列表，比如订单号
            lens = len(list_param)
            # 1 遍历special_param，每次1个
            print(f'根据{special_param}，每次获取1个数据')
            for i in tqdm(range(0, lens, 1)):
                # 2.2 获取对应页数的数据
                biz_content[special_param] = list_param[i: i + 1]
                pg_data = self.__get_data(biz_content, interface_method)
                if pg_data:
                    list_df.append(pd.DataFrame(pg_data))

        return list_df


class SQLServer:

    def __init__(self, server, user, password, database):
        self.conn = pymssql.connect(
            server=server,
            user=user,
            password=password,
            database=database,
            autocommit=True
        )
        self.cur = self.conn.cursor()

    def exec_query(self, sql):
        self.cur.execute(sql)
        self.conn.commit()

    def close(self):
        self.conn.close()


class Mysql:

    def __init__(self, host, user, password, database):
        self.conn = pymysql.connect(host=host, user=user, password=password, database=database, charset='utf8')
        self.cur = self.conn.cursor()

    def exec_query(self, sql):
        self.cur.execute(sql)
        self.conn.commit()
        result = self.cur.fetchall()
        for row in result:
            print(row)

    def close(self):
        self.conn.close()


def account_perriod(start_date=None, day='first') -> pd.Timestamp:
    """
    根据输入的月份格式返回日期格式
    :param start_date: 起始日期，如果有的话，返回这个日期对应的月的最后一天，比如传入'2023-01-01'，返回'2023-01-31'
    :param day: first、last，返回月初或者月末日期，默认为月初
    :return: "%Y-%m-01"
    """
    # 传入的日期格式固定为%Y-%m-%d

    # 判断有无输入日期
    if start_date:
        # 有输入日期，说明现在要求结尾日期，需要返回这个输入日期的当月最后一天。输入的日期加31天变成下个月日期，再减去天数变成当月最后一天
        default = (pd.to_datetime(start_date, format='%Y-%m-%d') + pd.Timedelta(31, 'D'))
        default = default - pd.Timedelta(default.day, 'D')
    else:
        # startDate为空，说明是求起始日期，默认上个月的第一天
        default = pd.Timestamp.now() - pd.Timedelta(pd.Timestamp.now().day + 1, 'D')  # 默认上个月
        default = default - pd.Timedelta(default.day - 1, 'D')

    month = input(f"格式为：2000-01，(直接回车就是这个账期)：{default.strftime('%Y-%m-%d')[:7]}\n")

    # 如果输入不为空
    if month:
        # 输入的格式是%Y-%m，先把它改成月份的第一天
        month = pd.to_datetime(month, format='%Y-%m')
        # 如果要返回月初，等于输入月份的第一天
        if day == 'first':
            date = month
        # 如果要返回月末，等于月份最后一天
        else:
            date = (month + pd.Timedelta(31, 'D'))
            date = (date - pd.Timedelta(date.day, 'D'))
    # 为空的话返回默认值
    else:
        date = default
    return date


def amz_used_item_regex() -> str:
    """
    从二手sku里获取sku的正则表达式
    """
    used_sku_regex = r'([A-Z]{1,2}\d?(-)?\d{0,2}[A-Z]{1,2}\d{2,6}[A-Z]{0,2})|(W\d{5})'
    return used_sku_regex


def any_files(folder_path: str) -> bool:
    """
    给文件地址，判断里面有没有文件，有的话返回True，反之False
    :param folder_path:
    :return:
    """
    for root, dirs, files in os.walk(folder_path):
        if files:
            return True
    return False


def cover_table(df, table_name, conn, server, user, password, database):
    """
    :feature: 清空table_name，添加df至《table_name》
    :return:
    """
    print(f'清空{table_name}')
    sql = 'delete from ' + table_name
    SQLServer(server, user, password, database).exec_query(sql)
    # 导入数据库
    print(f'导入至{table_name}')
    df.to_sql(table_name, conn, if_exists='append', index=False, chunksize=1000)
    print('导入完毕')


def create_dict_agg(df, merge_cols: list, func: str = 'sum'):
    """
    返回除了merge_cols之外其他列的聚合字典，默认sum
    :param df:
    :param merge_cols:
    :param func:
    :return:
    """
    dict_agg_temp = {}
    for col in df.columns:
        if col not in merge_cols:
            dict_agg_temp[col] = func
    return dict_agg_temp


def excel_process(file_path):
    """
    调整列宽，冻结首行，添加筛选
    freeze title, adjust width of columns, open filter
    :param file_path: path of file
    :return:
    """
    print('调整列宽，冻结首行，添加筛选')
    # 修改下述参数即可使用，Excel名称及Sheet名称即可
    work_book = openpyxl.load_workbook(file_path)
    for sheet in work_book.sheetnames:
        work_book[sheet].freeze_panes = 'A2'
        work_sheet = work_book[sheet]
        # 设置一个字典用于保存列宽数据
        dim_cols = {}
        # 遍历表格数据，获取自适应列宽数据
        for row in work_sheet.rows:
            for cell in row:
                if cell.value:
                    # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                    # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                    # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                    cell_len = 0.5 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                    dim_cols[cell.column] = max((dim_cols.get(cell.column, 0), cell_len))
        for col, value in dim_cols.items():
            # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的，限制最小宽度10， 最大宽度为30
            if value > 28:
                work_sheet.column_dim_colensions[get_column_letter(col)].width = 30
            elif value < 8:
                work_sheet.column_dim_colensions[get_column_letter(col)].width = 10
            else:
                work_sheet.column_dim_colensions[get_column_letter(col)].width = value + 2
        dict_num_to_alphabet = {
            1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J',
            11: 'K', 12: 'L', 13: 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T',
            21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y', 26: 'Z',
        }
        # 获取第一行
        rows = work_sheet.iter_rows(max_row=1, values_only=True)
        max_col = 0
        for cell in rows:
            max_col = len(cell)
        # 对第一行添加过滤功能
        filters = work_sheet.auto_filter
        filters.ref = 'A:' + dict_num_to_alphabet.get(max_col)
    work_book.save(file_path)
    print('end')
    print('$' * 20)


def group_by(df, by, func, dropna: bool = False, as_index: bool = False):
    """
    聚合
    param: func 可接受格式：
    '账期'
    ['账期', '花费']
    ['账期: max', '花费']
    默认sum
    """
    func_dict = {}
    if type(func) == str:
        func_dict[func] = 'sum'
    else:
        for unit in func:
            if len(re.split(': ', unit)) == 2:
                func_dict[re.split(': ', unit)[0]] = re.split(': ', unit)[1]
            else:
                func_dict[unit] = 'sum'

    return df.groupby(by, as_index=as_index, dropna=dropna).agg(func_dict)


def get_eu_sales_rate(
        file_path,
        include_uk=False,
        only_amz=True,
        include_platform=True,
        is_chuangyi=False
):
    """
    从业务数据中获取非英国的欧洲小组在事业部内销售额占比
    :param file_path: 业务数据的文件地址
    :param include_uk: 是否包含英国，默认不包含
    :param only_amz: 是否只取亚马逊数据，默认是
    :param include_platform: 是否包含平台维度，默认是
    :param is_chuangyi: 是否是创意的版本，默认否
    :return : [['平台', '小组', '事业部', 'rate']]
    """
    print('从业务数据中获取欧洲小组在事业部内非英国的销售额占比')

    if is_chuangyi:
        regex = '.*'
    else:
        regex = '.*-(DE|BE|FR|ES|IT|NL|PL|SE|TR)$'
        if include_uk:
            regex = regex[:-2] + '|UK' + regex[-2:]
        if only_amz:
            regex = '^AMAZON-' + regex

    # 第一次聚合的维度
    group_cols_1 = ['小组', '事业部', '账期']
    if include_platform:
        group_cols_1 += ['平台', '平台_true']

    # 第二次聚合的维度
    group_cols_2 = ['账期', '事业部']
    if include_platform:
        group_cols_2 += ['平台']

    # 小组过滤条件
    regex_group = '欧'
    if is_chuangyi:
        regex_group += '|创'

    df = pd.read_excel(file_path, usecols=['小组', '销售_原币', '兑美元汇率', '平台', '账期'])

    df = (
        df
        .loc[lambda d:
        d['平台'].str.contains(regex, regex=True, flags=re.IGNORECASE)
        & d['小组'].str.contains(regex_group, regex=True)
        & ~d['小组'].str.contains('综合')  # 去掉综合组
        ]
        .assign(
            sales=lambda d: d['销售_原币'] * d['兑美元汇率'],  # 原币销售额转成相同币种
            事业部=lambda d: d['小组'].str[:3],
            平台_true=lambda d: d['平台'],  # 实际的平台
            平台=lambda d: d['平台'].str[:-2] + 'DE',  # 欧洲平台都变成德国
        )
    )

    df = (
        df
        .groupby(group_cols_1, as_index=False)
        .agg({'sales': 'sum'})
        .assign(
            rate=lambda d: d.groupby(group_cols_2)['sales'].transform(lambda x: x / x.sum())
        )
        .astype({'账期': 'str'})
        .sort_values(group_cols_2)
        [group_cols_1 + ['rate']]
    )

    return df


def incremental_update(
        df: pd.DataFrame,
        table_name: str,
        merge_cols: list,
        conn: sqlalchemy.engine,
        server,
        user,
        password,
        database,
        dtype: dict = None,
        calculate_cols=None,
        dim_col: str = None
):
    """
    增量更新df至数据库的table_name中
    :param df: 需要被传入的DataFrame
    :param table_name: 需要更新的数据库表名
    :param merge_cols: 匹配的字段列表
    :param conn: 数据库连接
    :param server: 服务器地址
    :param user: 用户名
    :param password: 密码
    :param database: db名
    :param dtype: 特殊字段类型，默认空
    :param calculate_cols: 需要计算的字段，默认空，可接受字符串（单列）、列表（多列）
    :param dim_col: 维度列表，如果存在此参数，则不会根据merge_cols来判断本地与数据库的重复数据，而是通过此参数判断
    """
    # dtype设置为空白字典
    if dtype is None:
        dtype = {}

    # 获取当前系统用户名
    user_program = os.getlogin()
    # 重置索引，不然匹配时
    df = df.reset_index(drop=True)
    # 检测数据库中重复记录，有计算列的话会取计算列
    sql = f"select {', '.join(merge_cols)}" + ", id" if merge_cols else f"select {dim_col}" + ", id"
    if calculate_cols:
        if type(calculate_cols) == str:
            sql += f', {calculate_cols}'
        elif type(calculate_cols) == list:
            sql = sql + ', ' + ', '.join(calculate_cols)
    sql += f' from {table_name}'
    df_db = pd.read_sql(sql, conn, dtype=dtype)

    # 本地文件和数据库都有的数据
    if merge_cols:
        df_inner = df_db.merge(df, 'inner', merge_cols, )
    else:
        dim_tuple = tuple(df[dim_col].drop_duplicates().astype('str').values)  # 本地的dim_col去重后的维度元组
        df_inner = df_db.loc[lambda d: d[dim_col].isin(dim_tuple)]

    # 只有本地文件有的数据
    if merge_cols:
        df_not_inner = df.merge(df_inner[merge_cols].assign(mark=1), 'left', merge_cols, ).loc[
            lambda d: d['mark'].isnull()].drop('mark', axis=1)
    else:
        df_not_inner = pd.DataFrame({})

    nums_db = df_db.shape[0]  # 数据库的总数量
    nums_df = df.shape[0]  # 本地的总数量
    nums_inner = df_inner.shape[0]  # 两者交集的数量
    nums_only_df = df_not_inner.shape[0]  # 非交集的数量
    print('-' * 50)
    print(f'''数据库中“{table_name}”数据：{nums_db}条''')
    if dim_col:
        # 数据库的对应维度数量
        nums_db_dim = df_db.loc[lambda d: d[dim_col].isin(dim_tuple)].shape[0]
        print(f'''\t其中维度：{dim_col}：{dim_tuple}的：{nums_db_dim}条。''')
    print('-' * 50)
    print(f'''本地数据：{nums_df}条''')
    print('其中：') if nums_inner > 0 else None
    if nums_inner > 0:
        if dim_col:
            # 本地的对应维度数量
            nums_df_dim = df.loc[lambda d: d[dim_col].isin(dim_tuple)].shape[0]
            print(f'''\t其中维度：{dim_col}：{dim_tuple}的：{nums_df_dim}条。''')
        else:
            print(f'''\t两者都有的：{nums_inner}条''')
    print(f'''\t本地新的：{nums_only_df}条''') if nums_only_df > 0 else None

    # 有计算列的额外提示语句
    add_info = ''
    c1, c2 = '', ''
    if calculate_cols is not None:
        if type(calculate_cols) == str:
            # c1是数据库的金额，c2是本地文件的金额，如果有比较列，生成有差异的df
            c1, c2 = calculate_cols + '_x', calculate_cols + '_y'
            df_diff = df_db.merge(df, 'outer', merge_cols).loc[lambda d: (d[c1] != d[c2]) & (d[c2].notnull())]
            num_diff = df_diff.shape[0] - nums_inner
            add_info += f'两者都有的但“{calculate_cols}”数值不一致的：{num_diff}条' if num_diff > 0 else ''
        elif type(calculate_cols) == list:
            for calculate_col in calculate_cols:
                c1, c2 = calculate_col + '_x', calculate_col + '_y'
                df_diff = df_db.merge(df, 'outer', merge_cols).loc[lambda d: (d[c1] != d[c2]) & (d[c2].notnull())]
                num_diff = df_diff.shape[0]
                add_info += f'两者都有的但“{calculate_col}”数值不一致的：{num_diff}条\n\t' if num_diff > 0 else ''
    print(f'''\t{add_info}''') if add_info else None
    print('-' * 50)

    # 打印出重复信息
    if nums_inner > 5:
        df_print = df_inner.sample(1).reset_index()
    elif nums_inner == 0:
        df_print = pd.DataFrame({})
    else:
        df_print = df_inner.reset_index()
    for index in df_print.index:
        print('两者都有的随意一条数据：')
        print(df_print.loc[index])
        print('-' * 50)

    info = ('请输入指令以继续（直接回车可跳过）：'
            '\n1：清空数据库，再导入本地数据（此操作会清空历史数据，请慎重选择！）'
            '\n2: 只上传非重复记录（只添加新的）'
            '\n3: 删除数据不一致的，再上传非重复记录（更新老的，添加新的）'
            '\n\n')

    if add_info:
        info += '本地的数据跟数据库的相比，数值有变化，建议选：3'
    else:
        if nums_df == nums_inner:
            info += '本地的数据在数据库都有，建议回车跳过'
        elif nums_df == nums_only_df:
            info += '本地的数据在数据库都没有，建议选：2'
        else:
            info += '本地的数据有部分和数据库重合，建议选：3'
    info += '\n'
    select = input(info)

    # 要删除的id
    ids_to_be_deleted = df_inner['id'].to_list() if select == '3' else []
    # 要导入的数据，1和3都要导入所有本地数据
    df_upload = df_not_inner if select == '2' else df
    mysql = Mysql(server, user, password, database)

    # 先删除
    if select == '1':
        print(f'清空{table_name}')
        mysql.exec_query(f'truncate table {table_name}')
    elif select == '3':
        ids_len = len(ids_to_be_deleted)
        if ids_len > 0:
            # 删掉数据库中这些数据
            print(f'删除{table_name}中{ids_len}条记录')
            # 有维度的话按照维度删
            if dim_col:
                # 如果元组只有1个元素，把它变成字符串后替换掉逗号
                if len(dim_tuple) == 1:
                    dim_tuple = str(dim_tuple).replace(',', '')
                mysql.exec_query(f'delete {table_name} where {dim_col} in {dim_tuple}')
            else:
                for i in tqdm(range(0, ids_len, 1000)):
                    mysql.exec_query(f'delete from {table_name} where id in {tuple(ids_to_be_deleted[i: i + 1000])}')
            print(f'\n{user_program}删除了{table_name}的{ids_len}条记录')

    if select != '':
        # 导入数据库
        upload_records = df_upload.shape[0]
        # 再重置一次索引
        df_upload = df_upload.reset_index(drop=True)

        # 尝试添加update_time
        columns_db = pd.read_sql(f'select * from {table_name} limit 1', conn).columns.to_list()
        if 'update_time' in columns_db:
            print('添加update_time')
            df_upload['update_time'] = pd.Timestamp.now()

        print(f'{upload_records}条记录等待被导入至{table_name}')
        for i in tqdm(range(0, upload_records, 1000)):
            df_upload.loc[i: i + 999].to_sql(f'{table_name}', conn, index=False, if_exists='append')
        print(f'{user_program}导入了{upload_records}条记录至{table_name}')
    else:
        print('跳过')


def input_password(conn, mark: str = None):
    """输入密码，错误就退出程序"""
    # 返回小组列表
    df_password = pd.read_sql('''select password, account, regex from jybb_user_password''', conn)
    while True:
        # 返回小组列表，密码错误返回空列表
        password = input('请输入密码：')
        df_filtered = df_password.loc[lambda d: d['password'] == password][['account', 'regex',]]

        if df_filtered.size == 0:
            print('密码错误！请重新输入，如不清楚请联系信息部人员获得密码')
            continue
        else:
            account = df_filtered.values[0][0]
            print(f'您的小组是：{account}')
            return account


def lambda_f(n: str, sep: str = ','):
    """
    根据输入的关键词返回对应的匿名函数
    :param n: lam_multi_to_unique_single, s: lam_multi_to_single
    :param sep: 分隔符，默认逗号
    :return lambda function
    """
    lam_multi_to_unique_single = lambda x: re.sub('^,|,$', '', sep.join(set(x)))  # 多行转唯一一行
    lam_multi_to_single = lambda x: re.sub('^,|,$', '', sep.join(list(x)))  # 多行转一行
    return {'us': lam_multi_to_unique_single, 's': lam_multi_to_single}.get(n)


def order_files_exist(path) -> bool:
    """传入文件夹路径，判断里面的非“历史订单”文件夹中是否有文件，有的话返回True，否则返回False"""
    while True:
        for dir_path, _, file_list in os.walk(path):
            if dir_path[-4:] != '历史订单' and file_list != []:  # 订单文件夹里文件非空
                return True  # 有的话直接返回True
        return False  # 检测一遍后没有则返回False


def print_mapping_string(df: pd.DataFrame, duplicated_cols: list, key_word: str) -> None:
    """
    打印无映射的信息
    :param df: DataFrame
    :param duplicated_cols: 需要去重及打印的字段列表
    :param key_word: 关键词，告知什么东西没有映射
    """
    df = df.drop_duplicates(duplicated_cols).reset_index(drop=True)  # 去重后的df
    lens = df.shape[0]
    if lens > 0:
        print(f'以下记录无{key_word}：')
        # 打印要输出的文本
        for i in range(lens):
            # 循环拼接要打印的字符串
            word_list = []
            for k in range(len(duplicated_cols)):
                # df的第0列的第i行
                word_list.append(f'{duplicated_cols[k]}：{df[duplicated_cols[k]].loc[i]}')
            print('，'.join(word_list))


def remove_files():
    """移除除了历史文件以外的其他文件，订单、结算单之类"""
    file_list = []
    for path, dirName, fileNames in os.walk('../../raw_data'):
        for fileName in fileNames:
            # print(path)
            if not re.search('history', path):
                file_list.append(os.path.abspath(path + '\\' + fileName).replace('\\', '/'))

    if not file_list:
        print('无可删除文件')
    else:
        print('这些文件将被删除：')
        for j in file_list:
            # 展示绝对路径
            print(j)
        answer = input('情确认是否删除这些文件？（y/n）')
        if answer == 'y':
            for j in file_list:
                print('从file_list中删除%s' % j)
                os.remove(j)


def save_file(df, file_path, file_name, is_multi_df=False, dict_dfs=None):
    """
    经营报表第三步保存文件
    :param df:
    :param file_path:
    :param file_name:
    :param is_multi_df: is it multi df step, default False
    :param dict_dfs: map of workbook name to df, default null dict
    """
    file_path = file_path.replace('中间', '结果') + rf'\{file_name}.xlsx'
    # 判断是否是多df类型
    if is_multi_df:
        while True:
            try:
                with pd.ExcelWriter(file_path) as writer:
                    for workbook_name in dict_dfs:
                        print(f'写入{workbook_name}')
                        try:
                            dict_dfs[workbook_name].to_excel(writer, index=False, sheet_name=workbook_name)
                        except NotImplementedError:
                            dict_dfs[workbook_name].to_excel(writer, sheet_name=workbook_name)
                    break
            except PermissionError:
                input('文件打开了，请关闭文件后按回车继续保存')
                print('继续')
                continue

    else:
        if df.shape[0] > 0:
            print(f'保存文件至：{file_path}')
            while True:
                try:
                    df.to_excel(file_path, index=False)
                    print('导出完毕')
                    break
                except PermissionError:
                    input('文件打开了，请关闭文件后按回车继续保存')
                    print('继续')
                    continue
        else:
            print(f'{file_name}无有效数据，跳过导出')


def share_multisku_fees(
        df: pd.DataFrame,
        groupby_cols: list,
        sku_raw_col: str,
        calculate_cols: list,
        conn_152: sqlalchemy.engine
) -> pd.DataFrame:
    """
    用成本价来计算单品sku占组合的比例，继而分摊费用
    :param df:
    :param groupby_cols: 聚合维度，必须包含sku_raw_col
    :param sku_raw_col: sku原始列名，例如[品号-归属]
    :param calculate_cols: 费用列
    :param conn_152:
    """
    print('拆分组合sku，根据单品成本价在组合件内占比分摊费用')
    # 1 从数据库获取sku成本价
    mapping_cost = pd.read_sql(
        '''
        select sku, 出厂价, location
        from 参数2_成本头程关税
        ''', conn_152
    )

    # 2 生成agg的字典
    dict_agg = {}
    for col in calculate_cols:
        dict_agg[col] = 'sum'

    # 3 拆分组合件，根据成本价计算sku占比，分摊费用
    df = (
        df
        .groupby(groupby_cols, as_index=False, dropna=False)
        .agg(dict_agg)
        .assign(
            sku=lambda d: d[sku_raw_col].str.split('[+]')
        )
        .explode('sku')
        .assign(
            qty=lambda d: d['sku'].str.split('x').str[1].fillna(1).astype('int'),
            sku=lambda d: d['sku'].str.split('x').str[0],
            location=lambda d: np.where(d['sku'].str[-2:] == 'UK', 'uk', 'normal')
        )
        .merge(mapping_cost, 'left', ['sku', 'location'])
        .assign(
            # 出厂价填充空值为1，乘以数量
            出厂价=lambda d: d['出厂价'].fillna(1) * d['qty'],
            # 没有sku的话，要填充为1，不然费用没了
            rate=lambda d: d.groupby(groupby_cols)['出厂价'].transform(lambda x: x / x.sum()).fillna(1),
        )
    )

    # 4 要分摊的字段乘以比例
    for col in calculate_cols:
        df[col] = df[col] * df['rate']

    # 5 聚合
    groupby_cols.remove(sku_raw_col)  # 删掉原sku名字
    groupby_cols.append('sku')  # 加上sku作为维度
    df = (
        df
        # 删掉原来的sku名称，加上sku
        .groupby(groupby_cols, dropna=False)
        .agg(dict_agg)
    )

    return df


def unzip_file(zip_file_path, dest_path):
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        zip_ref.extractall(dest_path)


def update_exchange_rate(server, user, password, database):
    """
    复制最新汇率为上月汇率
    """
    sql = '''
        declare @last_month date, @current_month date
        set @last_month = cast(dateadd(mm, -1, dateadd(dd,-day(getdate())+1,getdate())) as date)
        set @current_month = cast(dateadd(dd,-day(getdate())+1,getdate()) as date)
        delete 汇率_月 where 月份 = @current_month --删除当月汇率
        insert into 汇率_月
        select 
        币种, 兑美元汇率, 月份=@current_month, 兑人民币汇率, 
        货币键=left(货币键, len(货币键)-6) + cast(year(@current_month)*100+month(@current_month) as varchar(6)), 
        国家, 排序 
        from 汇率_月 
        where 月份 = @last_month
    '''
    sql_server = SQLServer(server, user, password, database)
    sql_server.exec_query(sql)


def assign_division(df, col_name):
    """添加事业部字段"""
    df['事业部'] = np.select(
                    [
                        df[col_name].str.contains('AOK', flags=re.IGNORECASE),
                        df[col_name].str.contains('视'),
                        df[col_name].str.contains('办'),
                        df[col_name].str.contains('(家.+)|照', regex=True),
                        df[col_name].str.contains('创'),
                    ],
                    [
                        'AOK',
                        '视听',
                        '办公',
                        '家居',
                        '创意',
                    ],
                    '其他'
                )
    return df
