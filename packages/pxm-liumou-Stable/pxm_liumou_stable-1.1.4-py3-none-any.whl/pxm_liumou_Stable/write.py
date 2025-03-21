# -*- encoding: utf-8 -*-
"""
@File    :   write.py
@Time    :   2023-02-17 23:02
@Author  :   坐公交也用券
@Version :   1.0
@Contact :   faith01238@hotmail.com
@Homepage : https://liumou.site
@Desc    :   使用Openpyxl模块实现的写入功能
"""
from sys import exit

from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter

from .exists import _exists


class Write:
	def __init__(self, filename, create=True):
		"""
		初始化写入表格数据的类
		:param filename: 需要写入的表格文件路径
		:param create: 是否自动创建文件，默认为True
		"""
		# 文件不存在时是否自动创建
		self.auto_create_xlsx = create
		# 日志记录对象
		self.logger = logger
		# 表格文件名
		self.filename = filename
		# 检查文件是否存在，存在则加载，否则根据create参数决定是否创建
		if _exists(filename):
			self.logger.debug(f"使用加载文件的方式写入内容: {self.filename}")
			self.wb = load_workbook(filename=self.filename)
		else:
			if self.auto_create_xlsx:
				self.logger.debug(f"使用创建文件的方式写入内容: {self.filename}")
				self.wb = Workbook()
			else:
				self.logger.error(f"表格文件不存在且已关闭自动创建文件选项: {self.filename}")
				exit(2)
		# 获取工作表
		self.ws = self.wb.active
		# 是否自动添加序号，默认False
		self.SetSerialNumber = False
		# 是否设置添加，默认True
		self.SetAdd = True
		# 是否设置表头，默认True
		self.SetHead = True
		# 设置行数，默认None
		self.SetLine = None
		# 设置速度总和，默认100
		self.SetSpeedSum = 100
		# 需要写入的数据列表
		self._WrData = []
		# 错误信息，默认None
		self.Err = None

	def create_sheet(self, name, index=None):
		"""
		创建sheet
		:param name: 新sheet的名称 (必须为非空字符串)
		:param index: 设置sheet排序位置(0是排第一)，可选
		:return: 创建结果 (True 表示成功，False 表示失败)
		"""
		# 参数校验
		if not isinstance(name, str) or not name.strip():
			print("Error: Sheet name must be a non-empty string.")
			return False
		if index is not None and (not isinstance(index, int) or index < 0):
			print("Error: Index must be a non-negative integer or None.")
			return False

		try:
			# 创建sheet
			self.wb.create_sheet(title=name, index=index)
			return True  # 创建成功
		except Exception as e:
			print(f"Error: Failed to create sheet '{name}'. Details: {e}")
			return False  # 创建失败

	def save_workbook(self):
		"""
		保存工作簿
		"""
		try:
			self.wb.save(self.filename)
		except Exception as e:
			print(f"Error: Failed to save workbook. Details: {e}")

	def update_line(self, row, data_list):
		"""
		更新某行数据
		:param row: 需要更新的行
		:param data_list: 行数据列表
		:return: 成功返回 True，失败返回 False
		"""
		try:
			# 参数校验
			if not isinstance(row, int) or row < 1:
				raise ValueError(f"Invalid row index: {row}. Row index must be a positive integer.")
			if not isinstance(data_list, list):
				raise ValueError(f"Invalid data_list: {data_list}. Expected a list.")

			# 批量写入数据
			for col, value in enumerate(data_list, start=1):  # 列索引从1开始
				self.ws.cell(row=row, column=col, value=value)

			# 保存文件
			self.wb.save(self.filename)
			return True

		except ValueError as ve:
			# 记录参数校验错误
			self.logger.error(f"ValueError occurred while updating row {row}: {ve}")
			return False

		except Exception as e:
			# 记录其他异常并提供上下文信息
			self.logger.error(f"An error occurred while updating row {row} with data {data_list}: {e}")
			return False

	def write_cover_lists(self, lists, index=False):
		"""
		通过列表方式覆盖写入数据,一次性最多写入104万行数据
		:param lists: 写入数据列表,例如: [["张三", "男", "33"], ["李四", "男", "32"]]
		:param index: 是否保留首行
		:return:
		"""
		start_row = 1
		if index:
			start_row = 2
		try:
			# 计算要删除的行数
			max_row = self.ws.max_row
			if max_row >= start_row:
				rows_to_delete = max_row - start_row + 1
				self.ws.delete_rows(start_row, amount=rows_to_delete)
			# 写入新数据
			for row_data in lists:
				self.ws.append(row_data)
			# 保存到文件
			self.wb.save(self.filename)
			return True
		except Exception as err:
			self.Err = err
			self.logger.error(str(err))
			return False

	def write_lists(self, lists):
		"""
		通过列表方式追加写入数据,一次性最多写入104万行数据
		:param lists: 写入数据列表,例如: [["张三", "男", "33"], ["李四", "男", "32"]]
		:return: 成功返回True，失败返回False
		"""
		try:
			# 参数类型检查
			if not isinstance(lists, list):
				raise TypeError("参数 'lists' 必须是列表类型")
			if not all(isinstance(item, list) for item in lists):
				raise TypeError("列表中的每个元素必须是列表类型")

			# 数据量限制
			if len(lists) > 1040000:
				raise ValueError("数据量超过最大限制：104万行")

			# 批量写入（假设self.ws支持直接扩展列表）
			# 注意：此处需根据实际库的API调整，若append不支持批量，需保留原循环
			self.ws.extend(lists) if hasattr(self.ws, 'extend') else [self.ws.append(row) for row in lists]

			self.wb.save(self.filename)
			return True

		except (TypeError, ValueError) as err:  # 捕获类型和业务逻辑异常
			self.Err = str(err)
			self.logger.error(f"参数校验失败: {err}")
			return False
		except Exception as err:  # 捕获其他异常，记录详细信息
			self.Err = str(err)
			self.logger.error(f"写入失败: {err}", exc_info=True)
			return False

	def write_add_line(self, data):
		"""
		追加写入一行数据
		:param data: 数据,以列表形式 ["张三", "男", "33"]
		或者字典模式1: {"A": "刘某", "B": "男", "C": "22"}
		字典模式2: {1: 1, 2: 2, 3: 3}
		:return:
		"""
		try:
			# 处理字典类型数据
			if isinstance(data, dict):
				keys = list(data.keys())
				if all(isinstance(k, str) for k in keys):
					sorted_keys = sorted(keys)  # 按字母顺序排序
				elif all(isinstance(k, int) for k in keys):
					sorted_keys = sorted(keys)  # 按数值顺序排序
				else:
					raise ValueError("字典的键必须全部为字符串或整数")
				data = [data[k] for k in sorted_keys]
			elif not isinstance(data, list):
				raise TypeError("data必须是列表或符合要求的字典")

			# 执行追加操作
			self.ws.append(data)
			self.wb.save(self.filename)
			return True
		except (IOError, PermissionError) as err:
			self.Err = err
			self.logger.error(f"文件写入错误: {err}", exc_info=True)
			return False
		except Exception as err:
			self.Err = err
			self.logger.error(f"未知错误: {err}", exc_info=True)
			return False


	def write_add_col(self, col, data):
		"""
		将数据列表写入工作表的指定列

		:param col: 列标识符，支持字符串形式的列字母（如 'A'）或整数形式的列索引（如 1 表示第一列）
		:param data: 需要写入的字符串或数值类型的列表
		:return: bool类型，返回True表示写入成功，False表示发生错误

		函数流程说明：
			1. 验证并转换列参数格式
			2. 获取工作表指定列的所有单元格对象
			3. 将数据逐行写入对应单元格
			4. 保存工作簿并返回执行结果
		"""
		# 验证并转换列参数格式为标准列字母
		if isinstance(col, int):
			col = get_column_letter(col)
		elif not isinstance(col, str):
			raise TypeError("col 必须是字符串或整数")

		try:
			# 获取指定列的所有单元格对象
			column = self.ws[col]

			# 将数据逐行写入对应单元格（从第一行开始）
			for i in range(len(data)):
				column[i].value = data[i]

			# 保存修改后的工作簿文件
			self.ws.save(self.filename)
			return True

		except (IndexError, ValueError, IOError) as err:
			# 记录错误信息并返回失败状态
			self.Err = err
			self.logger.error(f"写入失败: {type(err).__name__} - {err}")
			return False


	def delete_line(self, index, row=1):
		"""
		删除指定起始行的连续多行数据

		参数:
		index (int): 需要删除的起始行号（最小值为1）
		row (int, optional): 需要删除的行数，默认值为1

		返回:
			self: 返回当前对象实例，支持链式调用

		异常处理:
			- 若参数类型或值无效，记录参数错误并设置错误信息
			- 其他未知错误将记录详细信息并设置错误信息
		"""
		self.Err = None
		try:
			# 执行参数有效性校验
			if not isinstance(index, int) or index < 1:
				raise ValueError("index 必须是大于等于1的整数")
			if not isinstance(row, int) or row < 1:
				raise ValueError("row 必须是大于等于1的整数")

			self.ws.delete_rows(idx=index, amount=row)
			self.wb.save(self.filename)
		except (ValueError, TypeError) as err:
			self.Err = f"参数错误: {err}"
			self.logger.error(f"删除行失败: index={index}, row={row}, 错误: {err}")
		except Exception as err:
			self.Err = f"未知错误: {err}"
			self.logger.error(f"删除行失败: index={index}, row={row}, 错误: {err}")
		return self



	def set(self, add: bool = True, head: bool = True, line: int = None, speed_sum: int = 100, serial_number: bool = False):
		"""
		设置写入参数
		:param add: 是否使用追加模式（默认值为True。当为True时追加写入，False时覆盖写入）
		:param head: 是否保留表头标题（默认值为True。当为True时保留首行标题，False时忽略标题）
		:param line: 自定义写入的行号（若需指定写入位置，传入整数行号；若为None则使用默认行为）
		:param speed_sum: 进度显示间隔（每写入指定行数后显示进度，必须为正整数）
		:param serial_number: 是否自动添加序号列（默认False。当为True时会在数据前添加连续序号）
		:return: None
		"""
		# 参数类型和有效性验证
		if line is not None and not isinstance(line, int):
			raise TypeError("line must be an integer or None")
		if not isinstance(speed_sum, int) or speed_sum <= 0:
			raise ValueError("speed_sum must be a positive integer")

		self.SetSerialNumber = serial_number
		self.SetAdd = add
		self.SetHead = head
		self.SetLine = line
		if not self.SetAdd:
			# 覆盖模式下强制设置初始行为0，覆盖用户传入的line值
			self.SetLine = 0
		self.SetSpeedSum = speed_sum



	def _center_whole(self):
		"""设置全局居中对齐所有单元格。

		Args:
		self (object): 类的实例对象。

		Returns:
		None
		"""
		logger.info("正在设置全局居中")
		align = Alignment(horizontal='center', vertical='center')
		try:
			# 获取工作表的行列范围并遍历所有单元格设置居中对齐
			min_row = self.ws.min_row
			max_row = self.ws.max_row
			min_col = self.ws.min_column
			max_col = self.ws.max_column
			for row in self.ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
				for cell in row:
					cell.alignment = align
		except Exception as e:
			# 记录错误信息并抛出异常
			logger.error(f"设置全局居中时发生错误: {e}")
			raise


	def _center_line(self, line: int):
		"""
		设置指定行的单元格内容水平居中对齐。
		:param line: 需要居中的行号（必须为大于等于1的整数）
		:return: None
		"""
		logger.info(f"正在设置[ {line} ]行居中")
		# 检查行号是否有效
		if line < 1:
			logger.error(f"无效的行号: {line}. 行号必须大于等于1")
			return
		# 尝试访问指定行，捕获索引越界或类型错误
		try:
			target_row = self.ws[line]
		except (IndexError, TypeError) as e:
			logger.error(f"访问行{line}时出错: {str(e)}")
			return
		# 检查目标行是否存在或为空
		if not target_row:
			logger.warning(f"行{line}不存在或为空，无法设置居中")
			return
		# 遍历行内所有单元格，保留原有对齐设置并设置水平居中
		for cell in target_row:
			current = cell.alignment or Alignment()
			new_alignment = current.copy(horizontal='center')
			cell.alignment = new_alignment


	def _center_col(self, col: int):
		"""
		设置指定列居中
		:param col: 需要居中的列（1-based）
		:return:
		"""
		align = Alignment(horizontal='center')
		# 使用 iter_cols 直接获取目标列的所有单元格
		for cell in self.ws.iter_cols(min_col=col, max_col=col, min_row=1, max_row=self.ws.max_row)[0]:
			cell.alignment = align


	def set_center(self, whole=True, line=None, col=None, lines=None, cols=None):
		"""
		设置居中,此功能会遍历所有参数进行居中设置(如果设置了全局则忽略后面的参数)
		:param whole: 全局居中(bool)
		:param line: 指定行居中(int)
		:param col: 指定列居中(int)
		:param lines: 指定行范围居中,通过列表传入需要居中的行([int, int])
		:param cols: 指定列范围居中,通过列表传入需要居中的列([int, int])
		:return:
		"""
		if whole:
			self._center_whole()
			return

		# 参数类型验证
		if lines is not None:
			if not isinstance(lines, list):
				raise TypeError("lines must be a list of integers")
			for row in lines:
				if not isinstance(row, int):
					raise TypeError("All elements in lines must be integers")
		if cols is not None:
			if not isinstance(cols, list):
				raise TypeError("cols must be a list of integers")
			for col_val in cols:
				if not isinstance(col_val, int):
					raise TypeError("All elements in cols must be integers")

		# 合并参数处理
		lines_list = []
		if line is not None:
			lines_list.append(line)
		if lines is not None:
			lines_list.extend(lines)

		cols_list = []
		if col is not None:
			cols_list.append(col)
		if cols is not None:
			cols_list.extend(cols)

		# 执行居中操作
		for line_num in lines_list:
			self._center_line(line=line_num)
		for col_num in cols_list:
			self._center_col(col=col_num)

