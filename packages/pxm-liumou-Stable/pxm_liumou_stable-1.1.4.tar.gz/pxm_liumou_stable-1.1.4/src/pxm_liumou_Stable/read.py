# -*- encoding: utf-8 -*-
"""
@File    :   read.py
@Time    :   2023-02-17 23:02
@Author  :   坐公交也用券
@Version :   1.0
@Contact :   faith01238@hotmail.com
@Homepage : https://liumou.site
@Desc    :   使用Openpyxl模块实现的读取功能
"""
from sys import exit

from loguru import logger
from openpyxl import load_workbook

from .dic import GetDic
from .exists import _exists


class Read:
	def __init__(self, filename):
		"""
		读取表格数据
		:param filename: 需要读取的表格文件
		"""
		self.logger = logger
		self.Err = None  # 设置错误信息

		self.debug = False
		self.filename = filename
		if not _exists(filename):
			self.logger.error(f"文件异常: {self.filename}")
			exit(1)
		# 实例
		self.wb = load_workbook(filename=self.filename)

		# 获取信息
		self.InfoSheetList = self.wb.sheetnames  # 获取所有sheet名称
		self.InfoSheetName = None  # 存储工作簿名称
		self.InfoSheet = None

		self.ws = self.wb.active
		self.InfoRows = self.ws.max_row  # 获取行数
		self.InfoCols = self.ws.max_column  # 获取列数

		# 设置数据变量
		self.DataR = []  # 首次读取的数据
		self.Data = []  # 最终处理的数据
		# 拆分数据
		self.DataSplit = []
		# 设置数据常量
		self._Dic = GetDic()
		self.set()  # 读取默认配置

	def set(self, debug=False, sheet_name=None, sheet_index=None):
		"""
		自定义设置
		:param debug: 是否开启Debug
		:param sheet_name: 通过工作簿名称选择,当设置了此项切勿再设置索引(默认通过索引设置/当同时设置则优先使用索引值)
		:param sheet_index: 通过工作簿索引值选择(默认:0)
		:return:
		"""
		self.debug = debug

		# 校验 InfoSheetList 是否为空
		if not self.InfoSheetList:
			logger.error("InfoSheetList 为空，无法设置工作簿")
			return False  # 返回错误状态，避免直接退出程序

		# 默认值初始化
		if sheet_name is None and sheet_index is None:
			self.InfoSheet = 0
		else:
			if sheet_index is not None and sheet_name is not None:
				# 同时设置了 sheet_name 和 sheet_index，优先使用 sheet_index
				self.InfoSheet = sheet_index
			elif sheet_index is not None:
				# 单独设置了 sheet_index
				if sheet_index < 0 or sheet_index >= len(self.InfoSheetList):
					logger.error(f"无效的 sheet_index: {sheet_index}")
					return False
				self.InfoSheet = sheet_index
			elif sheet_name is not None:
				# 单独设置了 sheet_name
				if sheet_name in self.InfoSheetList:
					self.InfoSheet = self.InfoSheetList.index(sheet_name)
				else:
					logger.error(f"找不到 sheet 名称: {sheet_name}")
					return False
			else:
				logger.error("参数设置错误，必须提供 sheet_name 或 sheet_index")
				return False

		# 设置 InfoSheetName
		try:
			self.InfoSheetName = self.InfoSheetList[self.InfoSheet]
		except IndexError:
			logger.error(f"索引超出范围: {self.InfoSheet}")
			return False

		# 指定工作簿
		try:
			self.wb.get_sheet_by_name(self.InfoSheetName)
		except Exception as e:
			logger.error(f"无法加载工作簿: {self.InfoSheetName}, 错误信息: {e}")
			return False

		return True  # 表示设置成功


	def base(self):
		"""
		打印表格文件基础信息
		:return:
		"""
		# 定义默认值，防止属性缺失导致异常
		filename = getattr(self, "filename", "未知文件")
		info_sheet_list = getattr(self, "InfoSheetList", [])
		info_sheet = getattr(self, "InfoSheet", "未选择sheet")
		info_rows = getattr(self, "InfoRows", 0)
		info_cols = getattr(self, "InfoCols", 0)

		# 使用info级别记录基础信息
		self.logger.info(f"当前文件: {filename}")
		self.logger.info(f"当前sheet列表: {info_sheet_list if info_sheet_list else '无'}")
		self.logger.info(f"当前选择sheet索引: {info_sheet}")
		self.logger.info(f"当前sheet总行数: {info_rows}")
		self.logger.info(f"当前sheet总列数: {info_cols}")


	def read_line(self, row=1):
		"""
		读取某行数据
		:param row: 需要读取的行
		:return: 包含该行所有列值的列表
		"""
		# 边界条件检查
		if not hasattr(self, 'ws') or self.ws is None:
			raise ValueError("工作表未初始化")
		if row < 1 or row > self.ws.max_row:
			raise IndexError(f"行号 {row} 超出工作表范围")

		data_ = []
		for col in range(1, self.InfoCols + 1):  # 直接从1开始计数
			try:
				value = self.ws.cell(row=row, column=col).value
				data_.append(value)
			except Exception as e:
				if self.debug:
					self.logger.error(f"读取单元格 ({row}, {col}) 时发生错误: {e}")
				data_.append(None)  # 如果读取失败，返回 None

		# 调试日志合并输出
		if self.debug:
			self.logger.debug(f"正在读取第 {row} 行数据: {data_}")

		return data_


	def read_col(self, col=1):
		"""
		读取某列所有数据
		:param col: 需要读取的列
		:return: 列数据列表
		"""
		# 参数校验
		if not isinstance(col, int) or col <= 0:
			raise ValueError(f"列号必须为正整数，当前值为: {col}")
		max_col = self.ws.max_column
		if col > max_col:
			raise ValueError(f"列号超出范围，最大列号为: {max_col}")

		data_ = []
		try:
			for row in range(1, self.InfoRows + 1):  # 调整为从 1 开始
				if self.debug:
					self.logger.debug(f"正在读取行: {row}, 列: {col}")
				value = self.ws.cell(row=row, column=col).value
				data_.append(value)
		except Exception as e:
			self.logger.error(f"读取列数据时发生错误: {e}")
			raise  # 将异常重新抛出以便调用方处理

		return data_


	def read_all(self, index=True):
		"""
		获取指定Sheet的所有数据。

		:param index: bool, 是否读取首行，默认为True表示读取首行。
		:return: self，便于链式调用。

		注意：
		- 如果index为True，则从第1行开始读取；否则从第2行开始读取。
		- 返回的数据存储在self.Data和self.DataR中。
		"""
		# 边界条件检查
		if self.InfoRows == 0 or self.InfoCols == 0:
			self.Data = []
			self.DataR = []
			return self  # 如果行列数为0，直接返回空数据

		data_ = []
		start_row = 1 if index else 2  # 根据index参数决定起始行

		try:
			for row in range(start_row, self.InfoRows + 1):  # 遍历行
				row_list = []
				for col in range(1, self.InfoCols + 1):  # 遍历列
					value = self.ws.cell(row=row, column=col).value
					row_list.append(value)
				data_.append(row_list)
		except Exception as e:
			# 异常处理，记录错误信息并返回空数据
			print(f"Error while reading data: {e}")
			self.Data = []
			self.DataR = []
			return self

		# 存储数据
		self.Data = data_ # 存储处理后的数据
		self.DataR = data_  # 存储原始数据,不会改变
		return self


	def _check(self, start, end, data):
		"""
		检查传入的数值是否符合实际要求
		:param start: 开始数
		:param end: 截止数
		:param data: 总数
		:return: (bool, str) 返回检查结果及错误信息
		"""
		try:
			# 将输入参数转换为整数
			start_int = int(start)
			end_int = int(end)
			data_int = int(data)

			# 检查负数情况
			if start_int < 0 or end_int < 0 or data_int < 0:
				return False, "输入参数不能为负数"

			# 检查开始数是否大于截止数
			if start_int >= end_int:
				return False, "开始数不能大于或等于截止数"

			# 检查截止数是否大于总数
			if end_int > data_int:
				return False, "截止数不能大于实际数"

			# 检查总数是否小于开始数
			if data_int <= start_int:
				return False, "实际数不能小于或等于开始数"

			# 如果所有条件都满足，返回 True
			return True, ""

		except ValueError:
			# 捕获类型转换异常，返回友好的错误信息
			return False, "输入参数必须是整数或可以转换为整数的字符串"


	def split_data(self, copies=1):
		"""
		对已处理的数据进行拆分
		:param copies: 需要拆分多少份,如果无法整除,那么最终份数会比设置的份数多1份
		:return: self.DataSplit
		"""
		# 检查输入参数的有效性
		if not isinstance(self.Data, list) or self.Data is None:
			logger.error("数据源无效，请确保 self.Data 是一个非空列表")
			return None
		if copies <= 0:
			logger.error(f"参数 copies 必须为正整数，当前值为 {copies}")
			return None

		# 初始化结果容器
		self.DataSplit = []

		# 获取总行数
		row = len(self.Data)
		if row == 0:
			logger.warning("数据为空，无需拆分")
			return self.DataSplit

		if copies == 1:
			logger.warning("未进行数据拆分...")
			self.DataSplit = [self.Data]
			return self.DataSplit

		# 计算每一份的数量和最后一份的数量
		number = row // copies
		number_end = row % copies

		# 核心拆分逻辑
		for i in range(copies):
			start_index = i * number
			end_index = (i + 1) * number
			if i == copies - 1:
				# 最后一份需要加上剩余的数据
				end_index += number_end
			self.DataSplit.append(self.Data[start_index:end_index])

		logger.info(f"根据参数对当前数据拆分成: {copies} 份, 每份数量: {number}, 其中最后一份数量: {number_end}")
		return self.DataSplit


	def read_line_range(self, start=1, end=1):
		"""
		读取行范围(1是第一行)
		:param start: 起始行 (默认为1，表示第一行)
		:param end: 结束行 (默认为1，表示第一行)
		:return: 获取结果(bool)，获取数据请通过实例变量(DataR)获取
		"""
		# 参数校验
		if not isinstance(start, int) or not isinstance(end, int):
			self.Err = "Invalid input: start and end must be integers."
			return False
		if start < 1 or end < 1:
			self.Err = "Invalid input: start and end must be positive integers."
			return False
		if start > end:
			self.Err = "Invalid input: start must be less than or equal to end."
			return False

		self.Err = None
		data_ = []
		try:
			# 确保范围正确
			row_start = start
			row_end = end + 1  # 调整范围以包含结束行

			if self._check(start=row_start, end=row_end, data=self.InfoRows):
				for row in range(row_start, row_end):
					row_list = []
					for col in range(1, self.InfoCols + 1):  # 列范围从1开始
						value = self.ws.cell(row=row, column=col).value
						row_list.append(value)
					data_.append(row_list)

				self.Data = data_
				self.DataR = data_

		except Exception as e:
			self.Err = f"Error occurred while reading lines: {str(e)}"
			return False

		return True


	def read_column_range(self, start=0, end=1):
		"""
		读取列范围
		:param start: 起始列
		:param end: 结束列
		:return: 获取结果(bool), 获取数据请通过实例变量(DataR)获取
		"""
		self.Err = None

		# 参数校验
		if not isinstance(start, int) or not isinstance(end, int):
			self.Err = "参数 'start' 和 'end' 必须为整数"
			return False
		if start < 0 or end < 0:
			self.Err = "'start' 和 'end' 不能为负数"
			return False
		if start >= end:
			self.Err = "'start' 必须小于 'end'"
			return False

		try:
			# 检查行数是否有效
			if not self._check(start=start, end=end, data=self.InfoRows):
				return False

			# 使用列表推导式优化数据读取
			data_ = [
				[self.ws.cell(row=row + 1, column=col).value for col in range(start, end)]
				for row in range(self.InfoRows)
			]

			# 合并数据赋值
			self.Data = self.DataR = data_

			return True  # 返回成功状态

		except Exception as e:
			self.Err = f"读取数据时发生错误: {e}"
			return False


	def cut_line(self, n=0):
		"""
		对已读取的数据进行截取指定行
		:param n: 需要截取的起始行号（非负整数）
		:return: 返回当前对象 (self)，操作结果可通过实例变量 (DataR) 获取
		"""
		# 检查 n 是否为非负整数
		if not isinstance(n, int) or n < 0:
			self.Err = "参数 n 必须是非负整数"
			self.logger.error(self.Err)
			return self

		# 确保 DataR 是列表类型
		if not isinstance(self.DataR, list):
			self.Err = f"DataR 类型错误，必须是列表，当前类型为 {type(self.DataR)}"
			self.logger.error(self.Err)
			return self

		# 检查是否存在未处理的错误
		if self.Err is not None:
			self.logger.error(f"已存在错误，请先处理: {self.Err}")
			return self

		# 检查数据长度是否足够
		if len(self.DataR) < n:
			self.Err = "行数据不足，无法截取"
			self.logger.error(self.Err)
			return self

		# 截取从第 n 行开始的数据
		self.DataR = self.DataR[n:]
		return self


	def cut_column(self, col=0):
		"""
		对已读取的数据进行列截取
		:param col: 需要截取的列
		:return: 返回 self 以支持链式调用
		"""
		# 检查输入参数是否合法
		if not isinstance(col, int):
			raise ValueError("参数 col 必须是整数")
		if col < 0:
			raise IndexError("参数 col 不允许为负数")

		# 检查数据源是否为空
		if not self.DataR:
			self.logger.warning("数据源为空，无法进行列截取")
			return self

		# 检查 logger 是否有效
		if not hasattr(self, 'logger') or self.logger is None:
			raise AttributeError("logger 未正确初始化")

		# 初始化结果列表
		data_list = []

		# 遍历数据并截取指定列
		for line, row in enumerate(self.DataR):
			if len(row) > col:  # 确保列索引有效
				data_list.append(row[col])
			else:
				self.logger.warning(f"第 {line} 行列数量不足，无法切割")
				self.logger.debug(row)

		# 更新实例变量
		self.DataR = data_list
		return self


	def cut_line_range(self, start=0, end=1):
		"""
		对已读取的数据进行 行范围 截取
		:param start: 起始行（非负整数）
		:param end: 结束行（非负整数）
		:return: 取结果(bool), 获取数据请通过实例变量(DataR)获取
		"""
		# 边界条件检查
		if not isinstance(start, int) or not isinstance(end, int) or start < 0 or end < 0:
			self._log_error("起始行或结束行必须是非负整数")
			return self
		if start > end:
			self._log_error(f"开始行数({start})大于结束行数({end})")
			return self

		# 异常处理：检查 DataR 是否为列表
		if not isinstance(self.DataR, list):
			self._log_error("DataR 数据类型错误，应为列表")
			return self

		# 检查是否已有错误
		if self.Err is not None:
			self._log_error(f"已存在错误, 请先处理: {self.Err}")
			return self

		# 检查数据长度是否满足截取需求
		if len(self.DataR) < end:
			self._log_error(f"行数据不足, 无法截取 (当前行数: {len(self.DataR)}, 截取范围: {start}-{end})")
			return self

		# 执行截取操作
		self.DataR = self.DataR[start:end]
		return self

	def _log_error(self, message):
		"""
		统一日志记录方法
		:param message: 错误信息
		"""
		self.Err = message
		self.logger.error(message)


	def cut_column_range(self, start=0, end=1):
		"""
		对已读取的数据进行 列范围 截取
		:param start: 起始列
		:param end: 结束列
		:return:
		"""
		# 边界条件检查
		if not isinstance(start, int) or not isinstance(end, int) or start < 0 or end < 0:
			self.logger.error("起始列和结束列必须是非负整数")
			self.Err = "起始列和结束列必须是非负整数"
			return self
		if start >= end:
			self.logger.error("起始列数大于或等于结束列数")
			self.Err = "起始列数大于或等于结束列数"
			return self

		# 检查 DataR 是否为可迭代对象
		if not hasattr(self.DataR, '__iter__'):
			self.logger.error("DataR 不是可迭代对象")
			self.Err = "DataR 不是可迭代对象"
			return self

		dataList = []  # 创建一个临时变量存储处理数据
		line = 0
		for i in self.DataR:
			try:
				# 检查当前行是否支持 len() 操作
				if not hasattr(i, '__len__'):
					raise TypeError(f"第 {line} 行数据不支持 len() 操作")

				# 检查列数是否足够
				if len(i) >= end:
					dataList.append(i[start:end])
				else:
					self.logger.warning(f"第 {line} 行列数量不足 (实际列数: {len(i)}), 无法切割")
					self.logger.debug(f"行内容: {i}")
			except Exception as e:
				self.logger.error(f"处理第 {line} 行时发生错误: {e}")
			finally:
				line += 1

		# 更新实例变量
		self.DataR = dataList
		return self
